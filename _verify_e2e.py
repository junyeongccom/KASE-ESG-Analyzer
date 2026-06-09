"""v3 / v4 실제 end-to-end 검증: 같은 CJ SR을 양쪽 경로로 돌려 출력 생성 확인.
앱이 호출하는 run_analysis_sync를 그대로 사용 (Gemini, 환경 시트만)."""
import sys, time
from pathlib import Path

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
sys.path.insert(0, str(BASE))
from dotenv import dotenv_values
import openpyxl
from core.providers import create_provider
from core.analyzer import run_analysis_sync

ENV = dotenv_values(str(BASE / ".env"))
PDF = Path("/Users/junyeongc/KASE/AI도입/03_원본보고서_PDF/2024_CJ제일제당_sustainability_report_ko.pdf").read_bytes()
prov = create_provider(provider_key="gemini", api_key=ENV["GOOGLE_API_KEY"], model="gemini-2.5-flash")
print(f"PDF {len(PDF)/1024/1024:.1f}MB, provider={prov.provider_name}")

CASES = [
    ("v4", BASE / "templates/식품_v4.xlsx", "식품_환경"),
]

for tag, tpl, sheet in CASES:
    print(f"\n===== {tag} : {sheet} =====")
    t = time.time()
    try:
        out, cost = run_analysis_sync(
            pdf_bytes=PDF, company_name=f"CJ_{tag}TEST",
            selected_sheets=[sheet], provider=prov, template_path=str(tpl))
    except Exception as e:
        print(f"  !! 실패: {type(e).__name__}: {e}")
        continue
    print(f"  완료 {time.time()-t:.0f}s | 비용 ${cost.get('total_cost',0):.4f} | -> {Path(out).name}")
    wb = openpyxl.load_workbook(out)
    ws = wb[sheet]
    hdr = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    name_col = hdr.get("Requirements") or hdr.get("지표명") or 3
    score_col = next((c for h, c in hdr.items() if "점수" in h), None)
    e_col = next((c for h, c in hdr.items() if "AS-IS" in h), None)
    filled = sum(1 for r in range(2, ws.max_row + 1)
                 if e_col and ws.cell(r, e_col).value and str(ws.cell(r, e_col).value).strip()
                 and "분석 대기" not in str(ws.cell(r, e_col).value))
    print(f"  출력열: 점수col={score_col} ASIScol={e_col} | 채워진 결과 {filled}행")
    shown = 0
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, name_col).value
        sc = ws.cell(r, score_col).value if score_col else None
        if nm and sc is not None and str(sc).strip() != "":
            print(f"    R{r} score={sc} | {str(nm)[:34]}")
            shown += 1
        if shown >= 4:
            break
    wb.close()

print("\nE2E 완료")

"""평가지표 템플릿 스키마 감지 + v4(YN 트리 + 메타데이터) 로더.

- v3: `Requirements`/`필수항목·배점0`/`AI AS-IS 내용` (기존 로직 = excel_handler.load_template)
- v4: `지표명`/`필수=N·Y→N·Y→Y→Y(만점)`/`Intent~Citation` (YN 트리 + 6 메타데이터)

스키마는 헤더 구조로 자동 판별한다(파일명에 의존하지 않음).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

# 헤더에 이 키워드가 보이면 해당 스키마로 판별 (소문자 비교)
_V4_MARKERS = ("intent", "지표명", "필수=n")
_V3_MARKERS = ("requirements", "ai as-is", "배점0")

# v4 메타데이터 6열 (1-indexed): M~R
_META_COLS = [
    ("intent", 13), ("terms", 14), ("accept", 15),
    ("reject", 16), ("keywords", 17), ("citation", 18),
]
_OUTPUT_HEADER = "AI AS-IS 내용"  # v4 출력열(재실행 시 기존 결과 감지용)


def detect_schema(template_path: str | Path) -> str:
    """첫 시트의 1행 헤더로 스키마를 판별한다. 기본값 'v3'."""
    wb = load_workbook(str(template_path), read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).lower() for c in row if c is not None]
            break
        blob = " | ".join(headers)
        if any(m in blob for m in _V4_MARKERS):
            return "v4"
        return "v3"
    finally:
        wb.close()


def _find_output_col(ws) -> int | None:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and _OUTPUT_HEADER in str(v):
            return c
    return None


def load_v4(template_path: str | Path) -> dict[str, list[dict]]:
    """v4 템플릿을 읽어 시트별 지표(YN 트리 + 메타데이터)를 반환한다."""
    wb = load_workbook(str(template_path), data_only=False)
    result: dict[str, list[dict]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out_col = _find_output_col(ws)
        indicators: list[dict] = []
        n = 0
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 3).value
            if not name or not str(name).strip():
                continue
            n += 1
            has_existing = False
            if out_col:
                ev = ws.cell(r, out_col).value
                has_existing = bool(ev and str(ev).strip())
            meta = {
                k: (str(ws.cell(r, c).value).strip() if ws.cell(r, c).value else "")
                for k, c in _META_COLS
            }
            try:
                max_score = float(ws.cell(r, 10).value) if ws.cell(r, 10).value not in (None, "") else 0
            except (TypeError, ValueError):
                max_score = 0
            indicators.append({
                "row": r,
                "indicator_number": f"({n})",
                "name": str(name).strip(),
                "must": ws.cell(r, 4).value,
                "s_mustN": ws.cell(r, 5).value,
                "add1": ws.cell(r, 6).value,
                "s_YN": ws.cell(r, 7).value,
                "add2": ws.cell(r, 8).value,
                "s_YYN": ws.cell(r, 9).value,
                "s_full": ws.cell(r, 10).value,
                "max_score": max_score,
                "formula": ws.cell(r, 11).value,
                "guide": ws.cell(r, 12).value,
                "meta": meta,
                "has_existing_content": has_existing,
            })
        result[sheet_name] = indicators
    wb.close()
    return result

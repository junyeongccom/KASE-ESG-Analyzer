"""엑셀 템플릿 읽기 / 결과 쓰기 모듈 — 헤더 기반 컬럼 자동 감지."""
from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

from config import (
    PLACEHOLDERS,
    TITLE_PATTERN_HARDCODED,
    TITLE_PATTERN_PLACEHOLDER,
    get_template_version,
)

# ── 헤더명 매핑 (여러 변형 지원) ──
# 키: 내부 필드명, 값: 헤더에서 찾을 키워드 목록
HEADER_KEYWORDS = {
    "indicator":    ["Requirements", "평가지표"],
    "category":     ["Category", "카테고리"],
    "topic":        ["Topic Disclosure", "Topic"],
    "required":     ["필수항목", "필수"],
    "score0":       ["배점0"],
    "bonus1":       ["가산항목1"],
    "score1":       ["배점1"],
    "bonus2":       ["가산항목2"],
    "score2":       ["배점2"],
    "guideline":    ["평가지침"],
    "note":         ["비고"],
    "max_score":    ["만점"],
    "criteria":     ["배점기준"],
    "e_content":    ["AI AS-IS 내용", "AI AS-IS"],
    "f_score":      ["AI 점수", "AI점수"],
    "g_review":     ["검토의견"],
}


def _detect_columns(ws) -> dict[str, int | None]:
    """Row 1의 헤더를 읽어 필드→컬럼번호 매핑을 반환한다."""
    col_map: dict[str, int | None] = {k: None for k in HEADER_KEYWORDS}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if not header:
            continue
        header_str = str(header).strip()
        for field, keywords in HEADER_KEYWORDS.items():
            for kw in keywords:
                if kw in header_str:
                    col_map[field] = col
                    break
    return col_map


def _detect_data_start_row(ws) -> int:
    """데이터 시작 행을 감지한다. Row 1이 헤더이면 2, Row 3이 헤더이면 4."""
    r1 = ws.cell(row=1, column=1).value
    # v5.2 스타일: Row 1이 타이틀(병합), Row 3이 헤더
    if r1 and ("SR" in str(r1) or "비교분석" in str(r1) or "[기업명]" in str(r1)):
        return 4
    # v3 스타일: Row 1이 컬럼 헤더
    return 2


def _is_placeholder(value: Any) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    if not s:
        return True
    return any(ph in s for ph in PLACEHOLDERS)


# ── 읽기 ──


def load_template(template_path: str | Path) -> dict[str, list[dict]]:
    """템플릿 엑셀을 읽어 시트별 지표 데이터를 반환한다. 헤더 기반 컬럼 자동 감지."""
    wb = load_workbook(str(template_path), data_only=False)
    result: dict[str, list[dict]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        col_map = _detect_columns(ws)
        start_row = _detect_data_start_row(ws)

        # 지표 컬럼 결정 (v3: "indicator", v5.2 fallback: column 2)
        ind_col = col_map.get("indicator") or 2
        e_col = col_map.get("e_content")
        indicators: list[dict] = []

        for row_idx in range(start_row, ws.max_row + 1):
            ind_val = ws.cell(row=row_idx, column=ind_col).value
            if not ind_val or not str(ind_val).strip():
                continue

            # 지표 번호 추출
            ind_text = str(ind_val).strip()

            # E열(AI AS-IS) 기존 내용 확인
            has_existing = False
            if e_col:
                e_val = ws.cell(row=row_idx, column=e_col).value
                has_existing = not _is_placeholder(e_val)

            # 배점 정보 구성
            criteria_text = ""
            max_score = 0

            # v3 구조: 필수/가산 분리
            if col_map.get("required"):
                parts = []
                req = ws.cell(row=row_idx, column=col_map["required"]).value
                s0 = ws.cell(row=row_idx, column=col_map["score0"]).value if col_map.get("score0") else 0
                if req:
                    parts.append(f"필수: {req} ({s0 or 0})")
                    max_score += float(s0 or 0)

                b1 = ws.cell(row=row_idx, column=col_map["bonus1"]).value if col_map.get("bonus1") else None
                s1 = ws.cell(row=row_idx, column=col_map["score1"]).value if col_map.get("score1") else 0
                if b1:
                    parts.append(f"가산1: {b1} ({s1 or 0})")
                    max_score += float(s1 or 0)

                b2 = ws.cell(row=row_idx, column=col_map["bonus2"]).value if col_map.get("bonus2") else None
                s2 = ws.cell(row=row_idx, column=col_map["score2"]).value if col_map.get("score2") else 0
                if b2:
                    parts.append(f"가산2: {b2} ({s2 or 0})")
                    max_score += float(s2 or 0)

                criteria_text = " / ".join(parts)

            # v5.2 구조: 배점기준 통합 컬럼
            elif col_map.get("criteria"):
                criteria_text = str(ws.cell(row=row_idx, column=col_map["criteria"]).value or "")

            # 만점
            if col_map.get("max_score"):
                ms = ws.cell(row=row_idx, column=col_map["max_score"]).value
                if ms:
                    max_score = float(ms)

            # 카테고리/토픽
            category = ""
            if col_map.get("category"):
                cat_val = ws.cell(row=row_idx, column=col_map["category"]).value
                if cat_val:
                    category = str(cat_val)
            if col_map.get("topic"):
                topic_val = ws.cell(row=row_idx, column=col_map["topic"]).value
                if topic_val:
                    category = f"{category} > {topic_val}" if category else str(topic_val)

            # 평가지침
            guideline = ""
            if col_map.get("guideline"):
                gl = ws.cell(row=row_idx, column=col_map["guideline"]).value
                if gl:
                    guideline = str(gl)

            indicators.append(
                {
                    "row": row_idx,
                    "indicator_number": "",  # 아래에서 채움
                    "category": category,
                    "indicator": ind_text,
                    "criteria": criteria_text,
                    "guideline": guideline,
                    "max_score": max_score,
                    "has_existing_content": has_existing,
                }
            )

        # 지표 번호 부여 (순서대로)
        for i, ind in enumerate(indicators, 1):
            ind["indicator_number"] = f"({i})"

        result[sheet_name] = indicators

    wb.close()
    return result


def get_sheet_names(template_path: str | Path) -> list[str]:
    wb = load_workbook(str(template_path), read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


# ── 쓰기 ──


def _ensure_output_file(
    template_path: str | Path,
    output_dir: str | Path,
    company_name: str,
) -> Path:
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    version = get_template_version(Path(template_path))
    filename = f"{company_name}_{version}_{today}_분석결과.xlsx"
    output_path = output_dir / filename
    if not output_path.exists():
        shutil.copy2(str(template_path), str(output_path))
    return output_path


def write_sheet_results(
    output_path: str | Path,
    company_name: str,
    sheet_name: str,
    results: list[dict],
) -> None:
    """분석 결과를 엑셀에 기입한다. 헤더 기반으로 쓰기 컬럼을 자동 감지."""
    wb = load_workbook(str(output_path))
    ws = wb[sheet_name]
    col_map = _detect_columns(ws)
    start_row = _detect_data_start_row(ws)

    e_col = col_map.get("e_content")
    f_col = col_map.get("f_score")
    g_col = col_map.get("g_review")

    if not e_col:
        wb.close()
        return  # 쓰기 컬럼이 없으면 스킵

    # Row 1 타이틀 기업명 교체 (v5.2 스타일)
    title_cell = ws.cell(row=1, column=1)
    if title_cell.value:
        title = str(title_cell.value)
        title = title.replace(TITLE_PATTERN_PLACEHOLDER, company_name)
        title = title.replace(TITLE_PATTERN_HARDCODED, company_name)
        title_cell.value = title

    # 지표 번호 → row 매핑
    ind_col = col_map.get("indicator") or 2
    row_map: dict[str, int] = {}
    counter = 0
    for row_idx in range(start_row, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=ind_col).value
        if val and str(val).strip():
            counter += 1
            row_map[f"({counter})"] = row_idx

    for item in results:
        ind_num = item.get("indicator_number", "")
        target_row = row_map.get(ind_num)
        if target_row is None:
            continue

        # 기존 실제 내용이 있으면 스킵
        if not _is_placeholder(ws.cell(row=target_row, column=e_col).value):
            continue

        wrap_align = Alignment(wrap_text=True, vertical="top")

        e_cell = ws.cell(row=target_row, column=e_col)
        e_cell.value = item.get("e_content", "")
        e_cell.alignment = wrap_align

        if f_col:
            try:
                ws.cell(row=target_row, column=f_col).value = float(item.get("f_score", 0))
            except (ValueError, TypeError):
                ws.cell(row=target_row, column=f_col).value = 0

        if g_col:
            g_cell = ws.cell(row=target_row, column=g_col)
            g_cell.value = item.get("g_review", "")
            g_cell.alignment = wrap_align

        # 행 높이 자동 조정 (E열 내용 기준, 1줄당 약 15pt)
        content = str(item.get("e_content", ""))
        col_width = 60  # E열 대략 너비 (문자 수)
        line_count = content.count("\n") + 1
        char_lines = max(1, len(content) // col_width)
        total_lines = max(line_count, char_lines)
        row_height = max(15, min(total_lines * 15, 300))  # 최소 15, 최대 300
        ws.row_dimensions[target_row].height = row_height

    wb.save(str(output_path))
    wb.close()


def prepare_output_file(
    template_path: str | Path,
    output_dir: str | Path,
    company_name: str,
) -> Path:
    return _ensure_output_file(template_path, output_dir, company_name)


# ── v4 결과 기록 ──

_V4_OUT_HEADERS = ["AI AS-IS 내용", "AI 점수", "검토의견"]
_V4_OUT_WIDTHS = {"AI AS-IS 내용": 60, "AI 점수": 8, "검토의견": 40}


def write_results_v4(
    output_path: str | Path,
    company_name: str,
    sheet_name: str,
    results: list[dict],
) -> None:
    """v4(YN 트리) 결과를 기록한다. 출력열이 없으면 시트 끝에 덧붙여 생성한다."""
    wb = load_workbook(str(output_path))
    ws = wb[sheet_name]

    # 기존 출력열 탐색
    colmap: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and str(v).strip() in _V4_OUT_HEADERS:
            colmap[str(v).strip()] = c

    # 없는 출력열은 끝에 생성 (네이비 헤더)
    navy = PatternFill("solid", start_color="16235A")
    white = Font(color="FFFFFF", bold=True)
    nxt = ws.max_column + 1
    for h in _V4_OUT_HEADERS:
        if h not in colmap:
            cell = ws.cell(1, nxt)
            cell.value = h
            cell.fill = navy
            cell.font = white
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(nxt)].width = _V4_OUT_WIDTHS[h]
            colmap[h] = nxt
            nxt += 1

    e_col, f_col, g_col = colmap["AI AS-IS 내용"], colmap["AI 점수"], colmap["검토의견"]

    # 지표 번호 → row 매핑 (지표명 C열 기준). 모델이 "(1)"/"1"/"지표 1" 등 어떤 형식으로
    # echo해도 매칭되도록 숫자만 추출해 비교한다.
    row_map: dict[str, int] = {}
    n = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 3).value
        if v and str(v).strip():
            n += 1
            row_map[str(n)] = r

    wrap = Alignment(wrap_text=True, vertical="top")
    for item in results:
        key = re.sub(r"\D", "", str(item.get("indicator_number", "")))
        target_row = row_map.get(key)
        if target_row is None:
            continue
        ec = ws.cell(target_row, e_col)
        ec.value = item.get("e_content", "")
        ec.alignment = wrap
        try:
            ws.cell(target_row, f_col).value = float(item.get("f_score", 0))
        except (ValueError, TypeError):
            ws.cell(target_row, f_col).value = 0
        gc = ws.cell(target_row, g_col)
        gc.value = item.get("g_review", "")
        gc.alignment = wrap

    wb.save(str(output_path))
    wb.close()

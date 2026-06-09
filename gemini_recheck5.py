"""공통오류 후보 5건을 Gemini로 fresh 재채점 (메타데이터 풀, 사전 점수 비공개).
→ 원래 합의(v3=v4) vs Claude(외부) vs Gemini재판단 3자 비교."""
import sys, re, json
from pathlib import Path

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
sys.path.insert(0, str(BASE))
from dotenv import dotenv_values
import openpyxl
from google import genai
from google.genai import types
from json_repair import repair_json
from core.schemas import load_v4
from verify_results import find_cols


def norm(s):
    return re.sub(r"[\s\W]", "", str(s or "")).lower()


ENV = dotenv_values(str(BASE / ".env"))
client = genai.Client(api_key=ENV["GOOGLE_API_KEY"])

EXT = json.loads((BASE / "output/external_judge_agreements.json").read_text())
NOISE = ("SS 배출", "할당량")  # Claude 라벨 노이즈 2건 제외
five = [x for x in EXT if x.get("verdict") == "wrong" and not any(n in x["name"] for n in NOISE)]

rubric = {norm(x["name"]): x for sn, inds in load_v4(BASE / "templates/식품_v4.xlsx").items() for x in inds}
V4X = next(BASE.glob("output/CJ_v4_v4_*분석결과.xlsx"))
wb = openpyxl.load_workbook(V4X, data_only=True)
ev = {}
for sn in ["식품_환경", "식품_사회"]:
    ws = wb[sn]
    nc, ec, _ = find_cols(ws)
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, nc).value
        if nm and str(nm).strip() and ec:
            ev[norm(nm)] = str(ws.cell(r, ec).value or "")
wb.close()

SYS = """당신은 ESG 평가 감리자입니다. 아래 지표를 '평가지표(채점기준 YN + 메타데이터: Intent/Accept/Reject)와 인용근거'에만 근거해 독립적으로 채점하세요.
- 일반지식으로 기준을 보완·완화하지 마세요. Accept/Reject 패턴을 엄격히 적용하세요.
- 가산조건은 근거에서 그 요건이 명시적으로 충족될 때만 인정하세요.
JSON만: {"results":[{"id":N,"score":숫자,"reason":"한 문장"}]}"""


def fmt(i, x):
    r = rubric.get(norm(x["name"]))
    if r:
        m = r.get("meta") or {}
        rub = (f"필수:{r['must']}(미충족 {r['s_mustN']}점)/가산1:{r['add1']}(필수만 {r['s_YN']}점)/"
               f"가산2:{r['add2']}(필수+가산1 {r['s_YYN']}점)/모두충족 {r['s_full']}점(만점)\n"
               f"  Intent:{m.get('intent','')}\n  Accept:{m.get('accept','')}\n  Reject:{m.get('reject','')}")
    else:
        rub = "(메타 없음)"
    return f"[id {i}] 지표:{x['name']}\n채점기준:{rub}\n인용근거:{ev.get(norm(x['name']), '')[:450]}"


byid = {}
for i, x in enumerate(five):  # 지표별 개별 호출 (부분응답 방지)
    resp = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[types.Content(parts=[types.Part.from_text(text=SYS + "\n\n" + fmt(i, x))])],
        config=types.GenerateContentConfig(max_output_tokens=2000, temperature=0.1, response_mime_type="application/json"))
    try:
        rr = json.loads(resp.text)
    except json.JSONDecodeError:
        rr = json.loads(repair_json(resp.text))
    items = rr.get("results") if isinstance(rr, dict) else rr
    if items:
        byid[i] = items[0]

def _lt(a, b):
    try:
        return float(a) < float(b)
    except (TypeError, ValueError):
        return False


out = []
print("═══ 5건 3자 비교: v3=v4(원래) / Claude(외부) / Gemini(재판단) ═══")
for i, x in enumerate(five):
    g = byid.get(i, {})
    gs = g.get("score")
    cons, cl = x["consensus"], x["claude_score"]
    if gs == cons and cl != cons:
        concl = "✅ 합의 정당 (Gemini=합의, Claude만 outlier)"
    elif gs == cl and gs != cons:
        concl = "🔴 공통 over-credit 유력 (Claude·Gemini 둘 다 일치)"
    elif _lt(gs, cons) and _lt(cl, cons):
        concl = "🟠 합의 과대 유력 (Claude·Gemini 둘 다 더 낮음)"
    else:
        concl = "⚪ 모호 (3자 상이 — 사람 판단)"
    out.append({"name": x["name"], "consensus": cons, "claude": cl, "gemini": gs,
                "gemini_reason": g.get("reason", ""), "conclusion": concl})
    print(f"\n• {x['name'][:44]}")
    print(f"   v3=v4={cons} │ Claude={cl} │ Gemini={gs}   {concl}")
    print(f"   Gemini 사유: {g.get('reason', '')[:100]}")

Path(BASE / "output/gemini_recheck5.json").write_text(json.dumps(out, ensure_ascii=False, indent=2, default=str))
print("\n저장: output/gemini_recheck5.json")

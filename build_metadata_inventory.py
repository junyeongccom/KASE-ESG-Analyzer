"""평가지표 v4 메타데이터(Intent/용어/Accept/Reject/키워드/Citation) 인벤토리 엑셀. 흩어진 지표별 규칙을 한눈에 + 특이규칙(Reject 보유) 플래그."""
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
sys.path.insert(0, str(BASE))
from core.schemas import load_v4
from structural_scorer import INDS, _ROUTE_NAMES, _norm_name  # v1.3.0 구조적채점 라우팅

_RID2TYPE = {i["id"]: i["type"] for i in INDS}
SCORED = {_norm_name(nm): f"구조적채점·{_RID2TYPE[rid]}" for rid, nm in _ROUTE_NAMES.items()}  # 정규화 지표명 → 채점방식

TPL = str(BASE / "templates/식품_v4.xlsx")
OUT = Path("/Users/junyeongc/KASE/AI도입/2_평가지표/평가지표_메타데이터_인벤토리.xlsx")
FONT = "맑은 고딕"
HEAD = PatternFill("solid", fgColor="305496")
HL = PatternFill("solid", fgColor="FFF2CC")        # Reject 셀 강조
STAR = PatternFill("solid", fgColor="FFEB9C")      # 특이규칙 행 플래그
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
META = [("intent", "의도(Intent)"), ("terms", "용어사전/동의어"), ("accept", "인정(Accept)"),
        ("reject", "불인정(Reject)"), ("keywords", "검색키워드"), ("citation", "출처요구(Citation)")]


def head(ws, headers, widths, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h); cell.font = Font(FONT, bold=True, color="FFFFFF"); cell.fill = HEAD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = THIN
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 26


td = load_v4(TPL)
recs = []
for sheet, inds in td.items():
    for ind in inds:
        m = ind.get("meta") or {}
        recs.append({"sheet": sheet.replace("식품_", ""), "num": ind["indicator_number"], "name": ind["name"],
                     **{k: (m.get(k) or "").strip() for k, _ in META}})

wb = openpyxl.Workbook()

# ── 탭1: 전체 메타데이터 카탈로그 ──
ws = wb.active; ws.title = "메타데이터 카탈로그"
ws["A1"] = "평가지표 v4 — 지표별 메타데이터 카탈로그 (★ = 불인정(Reject) 규칙 보유 = 특이규칙)"
ws["A1"].font = Font(FONT, size=12, bold=True)
cols = ["시트", "번호", "지표명", "★", "의도(Intent)", "용어사전", "인정(Accept)", "불인정(Reject)", "검색키워드", "출처요구", "v1.3.0 채점방식"]
head(ws, cols, [7, 6, 34, 4, 28, 22, 28, 30, 20, 16, 20], row=3)
SCFILL = PatternFill("solid", fgColor="C6EFCE")  # 구조적채점 셀 강조
r = 4
for rec in recs:
    star = "★" if rec["reject"] else ""
    scored = SCORED.get(_norm_name(rec["name"]), "")  # v1.3.0 구조적채점 라우팅 여부
    vals = [rec["sheet"], rec["num"], rec["name"], star, rec["intent"], rec["terms"],
            rec["accept"], rec["reject"], rec["keywords"], rec["citation"], scored]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(r, c, v); cell.font = Font(FONT, size=9); cell.border = THIN
        cell.alignment = Alignment(vertical="top", wrap_text=(c >= 5), horizontal=("center" if c in (1, 2, 4) else "left"))
    if rec["reject"]:
        ws.cell(r, 4).fill = STAR
        ws.cell(r, 8).fill = HL
    if scored:
        ws.cell(r, 11).fill = SCFILL
    ws.row_dimensions[r].height = 46
    r += 1
ws.freeze_panes = "C4"

# ── 탭2: 특이규칙 모음 (Reject 보유만) ──
ws2 = wb.create_sheet("특이규칙 모음")
flagged = [x for x in recs if x["reject"]]
ws2["A1"] = f"특이규칙 지표 {len(flagged)}개 — 불인정(Reject)/인정(Accept) 경계 규칙 한눈에"
ws2["A1"].font = Font(FONT, size=12, bold=True)
head(ws2, ["시트", "번호", "지표명", "불인정(Reject) — 이건 인정 안 함", "인정(Accept) — 이건 인정"], [7, 6, 34, 46, 40], row=3)
rr = 4
for x in flagged:
    vals = [x["sheet"], x["num"], x["name"], x["reject"], x["accept"]]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(rr, c, v); cell.font = Font(FONT, size=9); cell.border = THIN
        cell.alignment = Alignment(vertical="top", wrap_text=(c >= 3), horizontal=("center" if c in (1, 2) else "left"))
    ws2.cell(rr, 4).fill = HL
    ws2.row_dimensions[rr].height = 60
    rr += 1
ws2.freeze_panes = "A4"

# ── 탭3: 메타 충실도 요약 ──
ws3 = wb.create_sheet("메타 충실도")
ws3["A1"] = "메타데이터 필드별 충실도 (138개 지표 중 채워진 수)"; ws3["A1"].font = Font(FONT, size=12, bold=True)
head(ws3, ["메타 필드", "채워진 지표", "비율"], [22, 14, 10], row=3)
n_total = len(recs)
rr = 4
for k, label in META:
    n = sum(1 for x in recs if x[k])
    for c, v in enumerate([label, f"{n}/{n_total}", f"{n/n_total:.0%}"], 1):
        cell = ws3.cell(rr, c, v); cell.font = Font(FONT, size=10); cell.border = THIN
        cell.alignment = Alignment(horizontal=("left" if c == 1 else "center"))
    rr += 1
ws3.cell(rr + 1, 1, "※ Reject(불인정) 보유 지표 = 특이규칙 = 평가 경계가 까다로워 오류 잦은 후보군. 카탈로그 탭의 ★ 참고.").font = Font(FONT, size=9, italic=True, color="808080")

for s in wb.worksheets:
    s.sheet_view.topLeftCell = "A1"
wb.save(OUT)
print(f"✅ 생성: {OUT.name} ({len(wb.sheetnames)}탭) | 특이규칙(Reject 보유) {len(flagged)}개 / {n_total}개")

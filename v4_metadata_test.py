"""v4 메타데이터 A/B 실험.

질문: 환각났던 지표에 v4 메타데이터(Intent/Term/Accept/Reject/Keywords/Citation)를
넣으면 환각이 실제로 잡히는가?

설계: 동일한 개선 프롬프트(YN트리 점수로직 + 각주 강조 + 거부규칙)로
  A = 메타데이터 OFF
  B = 메타데이터 ON
를 Gemini로 같은 CJ SR에 돌려 비교. 대상 = 메타가 채워진(=과거 환각) 지표만.
메인 앱/excel_handler 안 건드리는 독립 하니스.
"""
from __future__ import annotations
import asyncio, json, time, sys
from pathlib import Path

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
sys.path.insert(0, str(BASE))
from dotenv import dotenv_values
import openpyxl
from core.providers.gemini_provider import GeminiProvider
from google.genai import types

ENV = dotenv_values(str(BASE / ".env"))
PDF_PATH = "/Users/junyeongc/KASE/AI도입/03_원본보고서_PDF/2024_CJ제일제당_sustainability_report_ko.pdf"
TEMPLATE = "/Users/junyeongc/KASE/AI도입/01_평가지표/KASE_평가지표_식품_v4.xlsx"
SHEET_TAGS = {"식품_환경": "환경", "식품_사회": "사회"}
META_COLS = [("intent", 13), ("terms", 14), ("accept", 15), ("reject", 16), ("keywords", 17), ("citation", 18)]


def load_targets():
    wb = openpyxl.load_workbook(TEMPLATE)
    out = {}
    for sn, tag in SHEET_TAGS.items():
        ws = wb[sn]
        rows = []
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 3).value
            if not name or not str(name).strip():
                continue
            meta = {k: ws.cell(r, c).value for k, c in META_COLS}
            if not any(v and str(v).strip() for v in meta.values()):
                continue  # 메타 없는 지표(=과거 환각 없던 것) skip
            rows.append({
                "id": f"{tag}-R{r}", "row": r, "name": str(name).strip(),
                "must": ws.cell(r, 4).value, "s_mustN": ws.cell(r, 5).value,
                "add1": ws.cell(r, 6).value, "s_YN": ws.cell(r, 7).value,
                "add2": ws.cell(r, 8).value, "s_YYN": ws.cell(r, 9).value,
                "s_full": ws.cell(r, 10).value,
                "formula": ws.cell(r, 11).value, "guide": ws.cell(r, 12).value,
                "meta": {k: (str(v).strip() if v else "") for k, v in meta.items()},
            })
        out[sn] = rows
    return out


SYSTEM = """당신은 ESG 지속가능경영보고서(SR) 평가 전문가입니다. 첨부된 SR PDF만을 근거로 각 지표를 평가하세요.

## 절대 규칙 (환각 방지)
- SR에 **그대로(verbatim) 인용 가능한 문장/수치**가 있을 때만 '공시됨'으로 인정한다.
- ESG 일반지식·업계평균·추측으로 빈칸을 메우지 마라. 근거 없으면 미충족으로 본다.
- **표 하단 주석과 작은 글씨 각주까지** 반드시 확인하라 (정의·수치가 각주에 있는 경우가 많다).
- 모든 인정 근거에는 [SR p.XX] 페이지 + 원문 인용을 붙여라. 근거 못 찾으면 "(공시확인불가)".

## 점수 = YN 의사결정 트리 (정확히 이 순서)
 1) 필수조건 미충족 → '필수=N 점수'
 2) 필수조건 충족, 가산조건1 미충족 → 'Y→N 점수'
 3) 필수+가산1 충족, 가산조건2 미충족 → 'Y→Y→N 점수'
 4) 필수+가산1+가산2 모두 충족 → '만점'
- 가산조건 칸이 '(없음)'/'(없음 → 바로 만점)'이면 그 분기는 없으며 직전 단계 충족이 최고점.
- ⚠️ "데이터를 찾음" ≠ "필수조건 충족". 조건 문구가 요구하는 바를 정확히 충족해야 한다.

## 출력 (JSON만)
- e_content는 **핵심 근거 원문·수치만 간결히** (최대 3줄). g_review는 1~2문장.
{"results":[{"indicator_number":"<제공된 id 그대로>","e_content":"근거 원문 [SR p.XX]","f_score":<숫자>,"g_review":"어느 분기로 몇 점인지 근거"}]}
"""


def fmt(ind, with_meta):
    L = [f"[지표 {ind['id']}] {ind['name']}",
         f"  · 필수조건: {ind['must']}  (미충족 시 {ind['s_mustN']}점)",
         f"  · 가산조건1: {ind['add1']}  (필수만 충족 시 {ind['s_YN']}점)",
         f"  · 가산조건2: {ind['add2']}  (필수+가산1 충족 시 {ind['s_YYN']}점)",
         f"  · 모두 충족 → {ind['s_full']}점 (만점)"]
    if ind['formula']:
        L.append(f"  · 참고산식: {ind['formula']}")
    if ind['guide']:
        L.append(f"  · 평가지침: {ind['guide']}")
    if with_meta:
        m = ind['meta']
        L += ["  [메타데이터]",
              f"   - 지표 의도(Intent): {m['intent']}",
              f"   - 용어 사전/동의어: {m['terms']}",
              f"   - 인정 형태(Accept): {m['accept']}",
              f"   - 불인정 반례(Reject): {m['reject']}",
              f"   - 검색 키워드: {m['keywords']}",
              f"   - 출처 요구(Citation): {m['citation']}"]
    return "\n".join(L)


def build_user(inds, with_meta):
    head = f"아래 {len(inds)}개 지표를 첨부 SR에서 평가하라. 각 id를 그대로 echo 하라.\n\n"
    return head + "\n\n".join(fmt(i, with_meta) for i in inds)


def _gemini(gem, pdf, sys_p, usr_p):
    resp = gem.client.models.generate_content(
        model=gem.model,
        contents=[types.Content(parts=[
            types.Part.from_bytes(data=pdf, mime_type="application/pdf"),
            types.Part.from_text(text=f"{sys_p}\n\n---\n\n{usr_p}")])],
        config=types.GenerateContentConfig(
            max_output_tokens=40000, temperature=0.1, response_mime_type="application/json"))
    m = resp.usage_metadata
    fin = str(resp.candidates[0].finish_reason) if resp.candidates else "?"
    usage = {"input_tokens": m.prompt_token_count if m else 0,
             "output_tokens": m.candidates_token_count if m else 0, "finish": fin}
    return resp.text, usage


async def run(gem, pdf, inds, with_meta):
    t = time.time()
    raw = ""
    try:
        raw, usage = _gemini(gem, pdf, SYSTEM, build_user(inds, with_meta))
        results = gem.parse_response(raw)
        return {"results": results, "usage": usage, "elapsed": time.time() - t,
                "error": None, "n": len(results), "finish": usage.get("finish")}
    except Exception as e:
        return {"results": [], "usage": {}, "elapsed": time.time() - t,
                "error": f"{type(e).__name__}: {e}", "n": 0, "finish": "?", "raw": raw[:400]}


async def main():
    targets = load_targets()
    pdf = Path(PDF_PATH).read_bytes()
    print(f"PDF {len(pdf)/1024/1024:.1f}MB | 대상: " +
          ", ".join(f"{k}={len(v)}" for k, v in targets.items()))
    gem = GeminiProvider(api_key=ENV["GOOGLE_API_KEY"], model="gemini-2.5-flash")

    allout, tin, tout = {}, 0, 0
    for sn, inds in targets.items():
        if not inds:
            continue
        print(f"\n{'#'*70}\n##### {sn}: {len(inds)}개 대상 지표 #####")
        a = await run(gem, pdf, inds, False)
        b = await run(gem, pdf, inds, True)
        for r in (a, b):
            tin += r["usage"].get("input_tokens", 0)
            tout += r["usage"].get("output_tokens", 0)
        print(f"  A(no-meta): {a['elapsed']:.0f}s n={a.get('n')} finish={a.get('finish')} err={a['error']}")
        print(f"  B(meta):    {b['elapsed']:.0f}s n={b.get('n')} finish={b.get('finish')} err={b['error']}")
        allout[sn] = {"indicators": inds, "A": a, "B": b}
        for i, ind in enumerate(inds):
            ra = a["results"][i] if i < len(a["results"]) else {}
            rb = b["results"][i] if i < len(b["results"]) else {}
            print(f"\n── {ind['id']}  {ind['name'][:46]}")
            print(f"   A[{ra.get('indicator_number','?')}] score={ra.get('f_score','?')} | {str(ra.get('e_content',''))[:150].replace(chr(10),' ')}")
            print(f"   B[{rb.get('indicator_number','?')}] score={rb.get('f_score','?')} | {str(rb.get('e_content',''))[:150].replace(chr(10),' ')}")

    cost = tin/1e6*0.30 + tout/1e6*2.50
    print(f"\n{'='*70}\n총 토큰: in={tin:,} out={tout:,} | 추정비용 ${cost:.4f} (Gemini 2.5 Flash)")
    Path(BASE / "v4_metadata_result.json").write_text(
        json.dumps(allout, ensure_ascii=False, indent=2, default=str))
    print("저장: v4_metadata_result.json")


asyncio.run(main())

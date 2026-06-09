"""v3·v4 검증 결과를 3탭 엑셀 한 파일로 정리.
  Tab1 v3_결과 (원본 그대로, 식품-E+S)
  Tab2 v4_결과 (원본 그대로, 식품_환경+사회)
  Tab3 v3v4_차이_LLM판정 (불일치 지점 + judge 점수·사유)"""
import json, glob
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = "/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer"
V3X = sorted(glob.glob(f"{BASE}/output/CJ_v3_v3_*분석결과.xlsx"))[-1]
V4X = sorted(glob.glob(f"{BASE}/output/CJ_v4_v4_*분석결과.xlsx"))[-1]
JUDGE = json.load(open(f"{BASE}/output/judge_result.json", encoding="utf-8"))
OUT = "/Users/junyeongc/Downloads/KASE_v3v4_검증_비교.xlsx"

KF = "맑은 고딕"
navy = PatternFill("solid", start_color="16235A")
white = Font(name=KF, color="FFFFFF", bold=True)
base = Font(name=KF)
wrap = Alignment(wrap_text=True, vertical="top")


def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill = navy
        cell.font = white
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"


def width_for(h):
    h = str(h or "")
    if any(k in h for k in ("AS-IS", "근거", "사유", "검토", "Intent", "Accept", "Reject", "Term", "평가지침")):
        return 55
    if "지표" in h or "Requirements" in h:
        return 42
    if "충돌" in h or "Citation" in h or "Search" in h or "결론" in h:
        return 38
    if "점수" in h:
        return 9
    return 16


def finalize(ws):
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = width_for(ws.cell(1, c).value)
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = base
            cell.alignment = wrap


def add_version_tab(title, path, sheets, ind_keys):
    ws = wb.create_sheet(title)
    src = load_workbook(path, data_only=True)
    hdr = None
    ind_c = None
    for sheet in sheets:
        s = src[sheet]
        if hdr is None:
            hdr = [s.cell(1, c).value for c in range(1, s.max_column + 1)]
            ind_c = next((c for c in range(1, s.max_column + 1)
                          if any(k in str(s.cell(1, c).value or "") for k in ind_keys)), 2)
            ws.append(["영역"] + [str(h) if h is not None else "" for h in hdr])
            style_header(ws, len(hdr) + 1)
        for r in range(2, s.max_row + 1):
            iv = s.cell(r, ind_c).value
            if not iv or not str(iv).strip():
                continue
            ws.append([sheet] + [s.cell(r, c).value for c in range(1, len(hdr) + 1)])
    src.close()
    finalize(ws)
    return ws.max_row - 1


import re


def _norm(s):
    return re.sub(r"[\s\W]", "", str(s or "")).lower()


wb = Workbook()
wb.remove(wb.active)
n3 = add_version_tab("v3_결과", V3X, ["식품-E", "식품-S"], ["Requirements", "지표명"])
n4 = add_version_tab("v4_결과", V4X, ["식품_환경", "식품_사회"], ["지표명", "Requirements"])

# v4_결과 탭 AI 결과 셀 색칠 (green=정확 / red=출처틀림 / orange=애매)
CAT = json.load(open(f"{BASE}/output/v4_categories.json", encoding="utf-8"))["map"]
COLORS = {"green": "C6EFCE", "red": "FFC7CE", "orange": "FFEB9C"}
_wsv4 = wb["v4_결과"]
_hdr = {str(_wsv4.cell(1, c).value): c for c in range(1, _wsv4.max_column + 1)}
_name_c = next(c for h, c in _hdr.items() if "지표명" in h)
_ai_cols = [c for h, c in _hdr.items() if ("AS-IS" in h or "AI 점수" in h or "검토의견" in h)]
for _r in range(2, _wsv4.max_row + 1):
    _col = CAT.get(_norm(_wsv4.cell(_r, _name_c).value))
    if _col in COLORS:
        _fill = PatternFill("solid", start_color=COLORS[_col])
        for _c in _ai_cols:
            _wsv4.cell(_r, _c).fill = _fill

# Tab3: 차이 + judge (+ 양쪽 실제 근거·검토의견)
import re


def _norm(s):
    return re.sub(r"[\s\W]", "", str(s or "")).lower()


def evidence_map(path, sheets, ind_keys):
    src = load_workbook(path, data_only=True)
    m = {}
    for sn in sheets:
        s = src[sn]
        hdr = {str(s.cell(1, c).value): c for c in range(1, s.max_column + 1) if s.cell(1, c).value}
        ic = next((c for h, c in hdr.items() if any(k in h for k in ind_keys)), 2)
        ec = next((c for h, c in hdr.items() if "AS-IS" in h), None)
        gc = next((c for h, c in hdr.items() if "검토" in h), None)
        for r in range(2, s.max_row + 1):
            nm = s.cell(r, ic).value
            if nm and str(nm).strip():
                m[_norm(nm)] = (str(s.cell(r, ec).value or "") if ec else "",
                                str(s.cell(r, gc).value or "") if gc else "")
    src.close()
    return m


ev3 = evidence_map(V3X, ["식품-E", "식품-S"], ["Requirements", "지표명"])
ev4 = evidence_map(V4X, ["식품_환경", "식품_사회"], ["지표명", "Requirements"])

ws = wb.create_sheet("v3v4_차이_LLM판정")
heads = ["지표명", "v3 점수", "v4 점수", "LLM 판정", "LLM 정답점수", "판정 사유(LLM)",
         "v3 근거(AS-IS)", "v3 검토의견", "v4 근거(AS-IS)", "v4 검토의견", "메타", "평가지표↔도메인 충돌"]
ws.append(heads)
style_header(ws, len(heads))
vmap = {"B": "✅ v4가 맞음", "A": "❌ v3가 맞음", "both_ok": "～ 둘 다 타당", "neither": "⚠ 둘 다 틀림"}
for x in JUDGE:
    k = _norm(x["name"])
    e3 = ev3.get(k, ("", ""))
    e4 = ev4.get(k, ("", ""))
    ws.append([x["name"], x["v3"], x["v4"], vmap.get(x["verdict"], x["verdict"]),
               x.get("correct"), x.get("reason", ""),
               e3[0], e3[1], e4[0], e4[1],
               "O" if x.get("has_meta") else "", x.get("conflict_note", "") if x.get("conflict") else ""])
finalize(ws)

# Tab4: 합의 외부검증 (Claude 96건 + Gemini 재판단 5건 + 3자 결론)
AGR = json.load(open(f"{BASE}/output/external_judge_agreements.json", encoding="utf-8"))
GR5 = json.load(open(f"{BASE}/output/gemini_recheck5.json", encoding="utf-8"))
gr5_by = {_norm(g["name"]): g for g in GR5}
vmap_cl = {"correct": "✅ 합의 확인", "wrong": "⚠ 합의 반박(공통오류 후보)", "uncertain": "? 불확실"}
order = {"wrong": 0, "uncertain": 1, "correct": 2}
ws = wb.create_sheet("합의_외부검증")
heads = ["지표명", "합의점수(v3=v4)", "Claude 판정", "Claude 점수", "Claude 사유", "Gemini 재판단", "3자 결론"]
ws.append(heads)
style_header(ws, len(heads))
for x in sorted(AGR, key=lambda a: order.get(a.get("verdict"), 3)):
    g = gr5_by.get(_norm(x["name"]), {})
    concl = g.get("conclusion") or ("ⓘ Claude 단독 반박 (Gemini 재검 제외 — 노이즈 가능)" if x.get("verdict") == "wrong" else "")
    ws.append([x["name"], x["consensus"], vmap_cl.get(x.get("verdict"), x.get("verdict")),
               x.get("claude_score"), x.get("reason", ""), g.get("gemini", ""), concl])
finalize(ws)

wb.save(OUT)
print("saved:", OUT)
print("tabs:", wb.sheetnames)

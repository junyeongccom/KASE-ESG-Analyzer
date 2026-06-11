"""
@file e2e_gate_v130.py
@description v1.3.0 e2e 게이트 — 실제 앱(run_analysis_sync) 경로로 6사 × 10회. 구조적 오버라이드된 결정론 11지표가 일관되게 나오는지(실제 시스템).
@module (검증 하니스)
@dependencies core.analyzer, core.schemas, core.providers, structural_scorer
"""
import sys, json, time, unicodedata
from pathlib import Path
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import dotenv_values
import openpyxl

sys.path.insert(0, "/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
from core.analyzer import run_analysis_sync
from core.schemas import load_v4
from core.providers import GeminiProvider
from structural_scorer import ROUTED, _norm_name, _ROUTE_NAMES, warm

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
PDFDIR = Path("/Users/junyeongc/KASE/AI도입/03_원본보고서_PDF")
TEMPLATE = str(BASE / "templates/식품_v4.xlsx")
ENV = dotenv_values(str(BASE / ".env"))
COMPANIES = [("CJ제일", "CJ제일제당"), ("풀무원", "풀무원"), ("오뚜기", "오뚜기"),
             ("daesang", "대상"), ("nestle", "네슬레"), ("unilever", "유니레버")]
RUNS = 10
WORKERS = 2  # 동시 실행 run 수 (rate limit·파일충돌 회피)
OUTDIR = Path("/Users/junyeongc/KASE/AI도입/02_분석결과/시스템검증_국내외_0610/_요약")
OUT = OUTDIR / "e2e_gate_v130.json"


def find_pdf(sub):
    for f in PDFDIR.glob("*.pdf"):
        if unicodedata.normalize("NFC", sub) in unicodedata.normalize("NFC", f.name):
            return f.read_bytes()
    raise FileNotFoundError(sub)


# 라우팅 11지표 → 템플릿 시트/번호 (selected_indicators 키 "시트|(n)")
tmpl = load_v4(TEMPLATE)
SHEETS = list(tmpl.keys())
selected = []
for sheet, inds in tmpl.items():
    for ind in inds:
        if _norm_name(ind["name"]) in ROUTED:
            selected.append(f"{sheet}|{ind['indicator_number']}")
assert len(selected) == 11, f"라우팅 매칭 실패! {len(selected)}/11"
print(f"[준비] 라우팅 11지표 selected_indicators OK | 시트={SHEETS}", flush=True)

PDFS = {coname: find_pdf(sub) for sub, coname in COMPANIES}
provider = GeminiProvider(api_key=ENV["GOOGLE_API_KEY"], model="gemini-2.5-flash")


def read_scores(xlsx_path):
    """출력 엑셀에서 라우팅 11지표의 'AI 점수'를 구조ID별로 읽는다."""
    wb = openpyxl.load_workbook(xlsx_path)
    out = {}
    for sheet in SHEETS:
        ws = wb[sheet]
        fcol = next((c for c in range(1, ws.max_column + 1)
                     if ws.cell(1, c).value and "AI 점수" in str(ws.cell(1, c).value)), None)
        if not fcol:
            continue
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, 3).value
            if v and str(v).strip():
                nn = _norm_name(v)
                if nn in ROUTED:
                    out[ROUTED[nn]] = ws.cell(r, fcol).value
    wb.close()
    return out


def one_run(coname, run_idx):
    path, _ = run_analysis_sync(PDFS[coname], f"{coname}_{run_idx}", SHEETS, provider,
                                TEMPLATE, None, selected)
    return coname, run_idx, read_scores(path)


def main():
    for b in PDFS.values():  # fitz 동시파싱 segfault 방지 — 6사 PDF 텍스트 캐시 단일스레드 예열
        warm(b)
    print(f"[준비] PDF 텍스트 캐시 예열 완료 ({len(PDFS)}사)", flush=True)
    tasks = [(co, r) for _, co in COMPANIES for r in range(RUNS)]
    res = {}
    done = 0
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=WORKERS) as ex:
        futs = [ex.submit(one_run, co, r) for co, r in tasks]
        for f in as_completed(futs):
            try:
                coname, run_idx, sc = f.result()
            except Exception as e:
                print(f"  ⚠ run 실패: {type(e).__name__}: {str(e)[:80]}", flush=True)
                continue
            for rid, v in sc.items():
                res.setdefault(coname, {}).setdefault(rid, []).append(v)
            done += 1
            print(f"  [{done}/{len(tasks)}] {coname}_{run_idx}: "
                  + " ".join(f"{rid.split()[0]}={sc.get(rid)}" for rid in _ROUTE_NAMES if rid in sc)
                  + f" | {time.time()-t0:.0f}s", flush=True)
            json.dump(res, open(OUT, "w"), ensure_ascii=False, default=str, indent=2)

    print("\n=== e2e 게이트 결과 (실제 앱 경로 · 구조적 오버라이드) ===", flush=True)
    total = con = 0
    for rid in _ROUTE_NAMES:
        cells = []
        for _, co in COMPANIES:
            sc = res.get(co, {}).get(rid, [])
            ok = len(sc) == RUNS and len(set(map(str, sc))) == 1
            cells.append("✅" if ok else f"❌{dict(Counter(map(str,sc)))}")
            total += 1
            con += ok
        print(f"  {rid:14}: " + " ".join(f"{co[:3]}{c}" for (_, co), c in zip(COMPANIES, cells)), flush=True)
    print(f"\n결정론 11지표 × 6사 = {total}셀 중 일관 {con}", flush=True)
    print(f"→ {'✅ e2e 게이트 통과' if con == total else '❌ 미통과'} | {OUT.name} 저장", flush=True)


if __name__ == "__main__":
    main()

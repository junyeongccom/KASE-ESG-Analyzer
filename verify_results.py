"""Phase 2 검증 하니스 (1단계: 무API).
결과 엑셀 + SR PDF를 받아 지표별로:
  ① 누락 검사: e_content/점수 채워졌나 (응답누락·API오류·빈칸 구분)
  ③ 출처 텍스트매칭: e_content의 핵심 수치가 SR 원문 텍스트에 실재하는가 (grounding 점수)
     → grounding 낮은 지표 = 비전 재검(2단계) 대상으로 표시

사용: python verify_results.py <result.xlsx>
출력: 콘솔 요약 + <result>_verify.json
"""
import sys, re, json
from pathlib import Path

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
sys.path.insert(0, str(BASE))
import fitz  # PyMuPDF
import openpyxl
from core.schemas import detect_schema

SR_PDF = "/Users/junyeongc/KASE/AI도입/03_원본보고서_PDF/2024_CJ제일제당_sustainability_report_ko.pdf"

NOEVIDENCE = ("공시확인불가",)
FAILMARK = ("응답 누락", "API 오류", "검토필요")
PLACEHOLDER = ("AI 분석 대기", "분석 전")  # 미분석(템플릿 기본값) 셀


def sr_norm_text():
    doc = fitz.open(SR_PDF)
    full = "\n".join(doc.load_page(i).get_text() for i in range(len(doc)))
    doc.close()
    return re.sub(r"[\s,]", "", full)  # 공백·콤마 제거 정규화


def find_cols(ws):
    name = e = sc = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "")
        if name is None and ("지표명" in h or "Requirements" in h):
            name = c
        if e is None and "AS-IS" in h:
            e = c
        if sc is None and "점수" in h:
            sc = c
    return name, e, sc


def distinctive_numbers(text):
    """연도(2019~2026) 제외, 유효숫자 2자리+ 또는 %/소수/콤마 포함 수치 토큰."""
    out = []
    for t in re.findall(r"\d[\d,]*\.?\d+%?|\d+%", text):
        clean = t.replace(",", "").rstrip("%")
        if not clean or "." not in t and re.fullmatch(r"20[12]\d", clean):
            continue
        digits = clean.replace(".", "")
        if len(digits) >= 2:
            out.append(t)
    return list(dict.fromkeys(out))  # 중복 제거


def grounding(e_content, sr):
    nums = distinctive_numbers(e_content)
    if not nums:
        return None, 0, 0  # 수치 근거 없음(서술형) → 텍스트매칭 불가, 비전 대상
    found = sum(1 for n in nums if n.replace(",", "").rstrip("%") in sr)
    return found / len(nums), found, len(nums)


COMMON = set("당사 회사 우리 있습니다 통해 위해 대한 관련 그리고 또한 모든 각각 해당 기준 현황 "
             "내용 결과 활동 운영 관리 추진 수립 확보 제고 강화 실시 다음 또는 경우 이를 통한 "
             "위한 대해 등을 등의 함께 그러나 따라 보고 공개 공시".split())


def kw_coverage(e_content, sr):
    """수치 외 서술형 근거: 고유 한국어 토큰(3자+)이 SR 원문에 실재하는 비율."""
    toks = [t for t in re.findall(r"[가-힣]{3,}", e_content) if t not in COMMON]
    toks = list(dict.fromkeys(toks))
    if not toks:
        return None, 0, 0
    found = sum(1 for t in toks if t in sr)
    return found / len(toks), found, len(toks)


def cited_pages(e_content):
    return sorted(set(int(x) for x in re.findall(r"[pP]\.?\s*(\d{1,3})", e_content or "")))


def classify(e_content, score):
    s = str(e_content or "").strip()
    if not s:
        return "빈칸"
    if any(m in s for m in PLACEHOLDER):
        return "미분석"
    if any(m in s for m in FAILMARK):
        return "누락(응답/오류)"
    if any(m in s for m in NOEVIDENCE):
        return "공시확인불가"
    return "정상"


def main(result_path):
    schema = detect_schema(result_path)
    sr = sr_norm_text()
    wb = openpyxl.load_workbook(result_path, data_only=True)
    report = {"file": Path(result_path).name, "schema": schema, "sheets": {}}
    print(f"\n===== 검증: {Path(result_path).name} (schema={schema}) =====")
    for sn in wb.sheetnames:
        ws = wb[sn]
        name_c, e_c, sc_c = find_cols(ws)
        if not (name_c and e_c):
            continue
        rows = []
        for r in range(2, ws.max_row + 1):
            nm = ws.cell(r, name_c).value
            if not nm or not str(nm).strip():
                continue
            ec = ws.cell(r, e_c).value
            score = ws.cell(r, sc_c).value if sc_c else None
            cat = classify(ec, score)
            ec_s = str(ec or "")
            g, gf, gn = grounding(ec_s, sr) if cat == "정상" else (None, 0, 0)
            kwc, kf, kn = kw_coverage(ec_s, sr) if cat == "정상" else (None, 0, 0)
            num_ok = g is not None and g >= 0.7
            kw_ok = kwc is not None and kwc >= 0.6
            needs_vision = cat == "정상" and not num_ok and not kw_ok
            rows.append({
                "row": r, "name": str(nm).strip(), "score": score, "cat": cat,
                "pages": cited_pages(ec_s),
                "grounding": round(g, 2) if g is not None else None, "gf": gf, "gn": gn,
                "kw_cov": round(kwc, 2) if kwc is not None else None, "kf": kf, "kn": kn,
                "needs_vision": needs_vision,
            })
        # 요약
        n = len(rows)
        cats = {}
        for x in rows:
            cats[x["cat"]] = cats.get(x["cat"], 0) + 1
        normal = [x for x in rows if x["cat"] == "정상"]
        with_num = [x for x in normal if x["grounding"] is not None]
        high = [x for x in with_num if x["grounding"] >= 0.7]
        vision = [x for x in rows if x["needs_vision"]]
        print(f"\n── {sn}: {n}지표")
        print(f"   ①누락검사: {cats}")
        if with_num:
            print(f"   ③출처(수치 텍스트매칭): {len(high)}/{len(with_num)} 高grounding(≥0.7) | 비전재검 대상 {len(vision)}")
        # 의심 상위 (낮은 grounding) 몇 개
        susp = sorted([x for x in with_num if x["grounding"] < 0.7], key=lambda x: x["grounding"])[:5]
        for x in susp:
            print(f"     ⚠️ R{x['row']} g={x['grounding']} ({x['gf']}/{x['gn']}) p{x['pages']} | {x['name']}")
        report["sheets"][sn] = {"n": n, "cats": cats,
                                "high_grounding": len(high), "with_num": len(with_num),
                                "needs_vision": len(vision), "rows": rows}
    wb.close()
    outp = Path(result_path).with_name(Path(result_path).stem + "_verify.json")
    outp.write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str))
    print(f"\n저장: {outp.name}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else None)

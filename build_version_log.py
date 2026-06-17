"""시스템 버전 로그 + 지표 규칙 레지스트리 엑셀 생성 (초기본). 이후 버전 bump는 이 xlsx를 직접 편집."""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("/Users/junyeongc/KASE/AI도입/SYSTEM_VERSION_LOG.xlsx")  # KASE 최상위(접근 쉬움, 코드 repo 밖)
OUT.parent.mkdir(exist_ok=True)
FONT = "맑은 고딕"
HEAD = PatternFill("solid", fgColor="305496")
GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
ORANGE = PatternFill("solid", fgColor="FFEB9C"); BANNER = PatternFill("solid", fgColor="D9E1F2")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def header(ws, headers, widths, row=1):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h); cell.font = Font(FONT, bold=True, color="FFFFFF"); cell.fill = HEAD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = THIN
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 28


def body(ws, rows, start, statuscol=None, wrapcols=()):
    for i, r in enumerate(rows):
        rr = start + i
        for c, v in enumerate(r, 1):
            cell = ws.cell(rr, c, v); cell.font = Font(FONT, size=10); cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=(c in wrapcols), horizontal=("center" if c in (1,) else "left"))
        if statuscol:
            s = str(ws.cell(rr, statuscol).value)
            ws.cell(rr, statuscol).fill = RED if "🔴" in s else (ORANGE if "🟡" in s else (GREEN if "🟢" in s else PatternFill()))
        ws.row_dimensions[rr].height = 42


wb = openpyxl.Workbook()

# ── 탭1: 안내 ──
ws = wb.active; ws.title = "안내"
ws["A1"] = "KASE ESG 평가 시스템 — 버전 로그 & 지표 규칙 레지스트리"
ws["A1"].font = Font(FONT, size=15, bold=True)
notes = [
    ("", False),
    ("평가지표(템플릿)는 v3/v4로 별도 버전관리됨 (templates/). 이 문서는 평가 시스템 = [시스템 프롬프트 + 평가 로직 + 모델 구성] 의 버전을 관리한다.", False),
    ("", False),
    ("■ 운영 규칙", True),
    ("  1. 시스템 프롬프트·로직·모델을 바꾸면 → 버전 bump + '버전 히스토리' 탭에 기록.", False),
    ("  2. 모든 변경엔 검증(before→after 지표) 첨부 — 재현성 / 출처검증율 / flash↔pro 일치율.", False),
    ("  3. 특정 지표 때문에 프롬프트 규칙을 넣으면 → '지표 규칙 레지스트리' 탭에 등록.", False),
    ("  4. ⚠️ 레지스트리에 등재된 지표(평가지표)를 수정할 땐, 연결된 시스템 프롬프트 규칙도 반드시 함께 검토.", False),
    ("", False),
    ("■ 지표별 규칙은 어디에 있나? (중요 — 2026-06-10 프롬프트 전체 감사)", True),
    ("  · 시스템 프롬프트(v3/v4) = 일반규칙만 (환각방지·YN트리·언어규칙). 특정 지표 규칙 0개 — 감사 확인.", False),
    ("  · 지표별 평가 가이드 = 평가지표 v4 템플릿의 메타데이터 6컬럼(Intent/용어사전/Accept/Reject/검색키워드/Citation)에 존재.", False),
    ("    → 지표와 '같은 행'이라 지표 수정 시 자동으로 함께 보임 = 별도 추적 불필요.", False),
    ("  · 따라서 이 레지스트리는 '시스템 프롬프트에 박는' 지표/클래스 규칙만 추적한다 (지표와 떨어져 있어 놓치기 쉬운 결합). v1.3.0부터 채워짐.", False),
    ("", False),
    ("■ 버전 규칙 (semver 유사) MAJOR.MINOR.PATCH", True),
    ("  · MAJOR : 스키마·평가 체계 자체 변경 (예: v3→v4 대응)", False),
    ("  · MINOR : 프롬프트 규칙·모델 구성 추가/변경 (평가 결과에 영향)", False),
    ("  · PATCH : 표기·뷰·버그 등 평가 결과에 영향 없는 수정", False),
    ("", False),
    ("■ 현재 시스템 버전 : v1.3.0   (평가지표 v4 / 평가자 flash + 결정론 11지표 구조적채점 / 판정자 gemini-3.1-pro-preview)", True),
    ("   코드 상수 core/prompt_builder.py 의 SYSTEM_VERSION 과 항상 일치시킬 것.", False),
    ("   ※ v1.1.0~v1.3.0 변경분 = main 반영 완료 (2026-06-12).", False),
]
for i, (t, b) in enumerate(notes, 2):
    ws.cell(i, 1, t).font = Font(FONT, size=11, bold=b)
ws.column_dimensions["A"].width = 115
ws.sheet_view.showGridLines = False

# ── 탭2: 버전 히스토리 ──
ws2 = wb.create_sheet("버전 히스토리")
header(ws2, ["버전", "날짜", "분류", "변경 내용", "사유", "검증 (before → after)"], [9, 12, 15, 40, 34, 38])
VH = [
    ["1.0.0", "2026-06-08", "MAJOR", "v4 스키마 라우팅(YN-tree 의사결정트리 + 6개 메타데이터) + Gemini flash 평가", "평가지표 v4(11컬럼) 도입", "— (baseline)"],
    ["1.1.0", "2026-06-10", "MINOR · 프롬프트", "언어 규칙: AS-IS 근거 = 보고서 원문 언어 축자 인용 / 검토의견 = 한국어 고정", "영문 SR(네슬레·유니레버)에서 한국어 의역 시 출처검증(축자일치) 실패 우려", "국외 출처검증 82~87% = 국내(84%) 범위 → 글로벌 정상평가 확인"],
    ["1.2.0", "2026-06-10", "MINOR · 인프라", "독립 판정자 gemini-3.1-pro-preview(다른 세대=탈상관) + 모델별 조건부 thinking(flash off/판정 on) + 2.5-pro 폴백", "평가(flash) 교차검증용 독립 판정자 필요", "판정 재현성 95% (131/138, flash 67% 대비)"],
    ["1.3.0", "2026-06-12", "MINOR · 로직", "결정론 11지표 구조적 채점기 도입 — LLM '점수 판단' → 코드가 PDF 교차검증 후 결정론 채점 (core/analyzer._apply_structural_override). 연도 날조·수치 자기모순·각주 무시·취수≠조달 혼동 제거", "stably-wrong(재현O·정확X) 14개 중 11개 = flash가 일관되게 오답 → 판단을 코드로 대체", "전 식품 6사×10회 e2e(실제 앱): 결정론 11지표×6사=66셀 전부 10/10 일관 + 기준기업 정답 11/11=pro 일치"],
]
body(ws2, VH, 2, wrapcols=(4, 5, 6))
ws2.freeze_panes = "A2"

# ── 탭3: 지표 규칙 레지스트리 ──
ws3 = wb.create_sheet("지표 규칙 레지스트리")
ws3["A1"] = "⚠️ 이 표의 지표(평가지표 템플릿)를 수정할 땐, '시스템 프롬프트 규칙' 칸도 반드시 함께 검토할 것."
ws3["A1"].font = Font(FONT, size=10, bold=True, color="C00000")
header(ws3, ["#", "지표명", "시트", "상태", "발견된 문제 (테스트 근거)", "적용 규칙 (구조적채점·코드)", "버전", "검증"], [4, 34, 7, 11, 40, 40, 8, 11], row=2)
RG = [
    [1, "최근 3개년 산업재해율(%)을 공개하고 있다", "사회", "🟢 해결", "flash가 2개년(2023·24) 표를 2022 컬럼 날조해 '3개년' 만점 (풀무원 p.148 확정)", "구조적채점(year_count): LLM은 공시연도만 추출→코드가 PDF 교차검증·카운트, N개 미만이면 0", "1.3.0", "✅ 66/66"],
    [2, "최근 3개년 근로손실재해율(LTIFR)을 공개하고 있다", "사회", "🟢 해결", "동일 — 2개년인데 3개년 날조", "구조적채점(year_count)", "1.3.0", "✅ 66/66"],
    [3, "최근 3개년 협력업체 산업재해율(%)을 공개하고 있다", "사회", "🟢 해결", "동일", "구조적채점(year_count)", "1.3.0", "✅ 66/66"],
    [4, "전체 임직원 대비 정규직 임직원 비율이 80% 이상이다", "사회", "🟢 해결", "정규직 73.3%(<80%)인데 만점 — 공시와 조건충족 혼동", "구조적채점(threshold): 정규직수·전체수 추출→코드가 비율 계산·기준 비교", "1.3.0", "✅ 66/66"],
    [5, "최근 3개년 총폐기물 대비 지정폐기물(매립+소각) 비율 개선", "환경", "🟢 해결", "엉뚱한 지표(재활용률) 인용", "구조적채점(computed_trend): 매립·소각·총 추출→코드가 산식 계산·추세 판정", "1.3.0", "✅ 66/66"],
    [6, "(먼지) 최근 3개년 대기오염물질 배출 실적 개선", "환경", "🟢 해결", "필수(당해 감소)↔3개년 연속 혼동 (pro 승 확정)", "구조적채점(trend): 연도별 값 추출→코드가 당해감소 판정", "1.3.0", "✅ 66/66"],
    [7, "(황산화물) 최근 3개년 대기오염물질 배출 실적 개선", "환경", "🟡 보류", "보고경계(단독 vs 합계). 관계사 2024부터 측정→합계 추세 왜곡", "KASE 보고경계 정책 확정 후 구조적채점(trend) 적용", "(대기)", "➖ 정책"],
    [8, "물 스트레스 지역 조달 식품 원료 비율 정량 공개", "환경", "🟢 해결", "취수량(withdrawn)≠조달비율 / SASB 인덱스 라벨('-')을 공시로 오인", "구조적채점(presence_code): 물스트레스 근접 소싱어+실제 소수%값 코드검출 (풀무원만 공시=2)", "1.3.0", "✅ 66/66"],
    [9, "임시직 수 또는 비율을 공개하고 있다", "사회", "🟢 해결", "각주 제외조건(일용직·노무파견·도급) 무시→완전성 만점", "구조적채점(completeness): 공시표현·제외표현 코드검출→완전성 감점(1.5)", "1.3.0", "✅ 66/66"],
    [10, "정보 관련 사고 처리 대응 체계를 마련하고 있다", "사회", "🟡 보류", "지표 정의 미흡 — '사고 대응체계'를 flash가 '유출 진단'으로 확대해석 (rubric)", "평가지표 구체화 (메타 Accept/Reject: 진단·모니터링 ≠ 사고 대응체계)", "(대기)", "➖ 평가지표"],
    [11, "제품·서비스 안전성 제3자 인증 검증 내역 공개", "사회", "🟡 보류", "cruelty-free/윤리인증이 '안전성' 검증인지 정의 미흡 (rubric)", "평가지표 구체화 (안전성 검증에 윤리인증 포함 여부 정의)", "(대기)", "➖ 평가지표"],
    [12, "최근 3개년 온실가스 Scope 3 배출량 공개", "환경", "🟢 해결", "2개년인데 3개년 날조", "구조적채점(year_count)", "1.3.0", "✅ 66/66"],
    [13, "최근 3개년 온실가스 배출 집약도 공개", "환경", "🟢 해결", "2개년인데 3개년 날조", "구조적채점(year_count)", "1.3.0", "✅ 66/66"],
    [14, "최근 3개년도 불공정거래/부정경쟁 제재 내역 공개", "사회", "🟢 해결", "2개년 표를 3개년 날조", "구조적채점(year_count)", "1.3.0", "✅ 66/66"],
]
body(ws3, RG, 3, statuscol=4, wrapcols=(2, 5, 6))
ws3.cell(18, 1, "🟢 해결 11개 = 구조적채점(v1.3.0, 전 식품 6사×10회 e2e 66/66 일관 + 기준기업 정답 11/11). 🟡 보류 3개 = #7 황산화물(KASE 보고경계 정책 결정 대기)·#10 정보사고·#11 제3자인증(평가지표 루브릭 구체화 대기).").font = Font(FONT, size=9, italic=True, color="808080")
ws3.freeze_panes = "A3"

# ── 탭4: 검증 기준선 ──
ws4 = wb.create_sheet("검증 기준선")
ws4["A1"] = "검증 기준선 (v1.3.0)"; ws4["A1"].font = Font(FONT, size=12, bold=True)
header(ws4, ["지표", "값", "비고"], [36, 20, 40], row=3)
KPI = [
    ["flash 재현성 (전사 안정)", "26 / 138 (19%)", "v1.2.0 기준 — 6사 모두 10회 동일"],
    ["★ 진짜 신뢰 (재현 ∩ 정확)", "12 → +11 (구조적채점)", "핵심 KPI. stably-wrong 11개를 구조적채점으로 정확성 회복"],
    ["stably-wrong (재현 O·정확 X) 14개", "11 해결 / 3 보류", "11=구조적채점(v1.3.0) · 3=정책·평가지표 트랙"],
    ["★ v1.3.0 e2e 게이트 (실제 앱)", "66 / 66 일관", "결정론 11지표×6사 10/10 + 기준기업 정답 11/11=pro"],
    ["국내 / 국외 출처검증", "84% / 82~87%", "동등 (글로벌 정상평가)"],
]
body(ws4, KPI, 4, wrapcols=(3,))
ws4.cell(5, 1).font = Font(FONT, size=10, bold=True); ws4.cell(5, 1).fill = GREEN
ws4.cell(7, 1).font = Font(FONT, size=10, bold=True); ws4.cell(7, 1).fill = GREEN
ws4.cell(11, 1, "산출물: 4_결과/v1.3.0_structural/ (structural_repro_allco.json · e2e_gate_v130.json · v1.3.0_개선과정.xlsx) · 4_결과/시스템검증_국내외_0610/_요약/KASE_v4_검증종합_0610.xlsx").font = Font(FONT, size=9, italic=True, color="808080")

# 모든 시트 맨 위에서 열리게
for s in wb.worksheets:
    s.sheet_view.topLeftCell = "A1"

wb.save(OUT)
print(f"✅ 생성: {OUT.name} ({len(wb.sheetnames)}탭: {wb.sheetnames})")

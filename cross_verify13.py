"""미교차검증 13건(S 미매칭·출처 OK)을 외부(Claude)로 점수 검증 → cross13.json."""
import sys, re, json, time
from pathlib import Path

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
sys.path.insert(0, str(BASE))
from dotenv import dotenv_values
import openpyxl
import anthropic
from json_repair import repair_json
from core.schemas import load_v4
from verify_results import find_cols


def norm(s):
    return re.sub(r"[\s\W]", "", str(s or "")).lower()


ENV = dotenv_values(str(BASE / ".env"))
client = anthropic.Anthropic(api_key=ENV["ANTHROPIC_API_KEY"])
MODEL = "claude-sonnet-4-20250514"

V4X = next(BASE.glob("output/CJ_v4_v4_*분석결과.xlsx"))
V4J = json.loads((BASE / "output" / (V4X.stem + "_verify.json")).read_text())
dis_by = {norm(x["name"]) for x in json.loads((BASE / "output/judge_result.json").read_text())}
agr_by = {norm(x["name"]) for x in json.loads((BASE / "output/external_judge_agreements.json").read_text())}
rubric = {norm(x["name"]): x for sn, inds in load_v4(BASE / "templates/식품_v4.xlsx").items() for x in inds}

wb = openpyxl.load_workbook(V4X, data_only=True)
ev = {}
for sn in ["식품_환경", "식품_사회"]:
    ws = wb[sn]
    nc, ec, _ = find_cols(ws)
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, nc).value
        if nm and str(nm).strip() and ec:
            ev[norm(nm)] = str(ws.cell(r, ec).value or "")
wb.close()

D = []
for sn, data in V4J["sheets"].items():
    for x in data["rows"]:
        k = norm(x["name"])
        if x["cat"] == "정상" and not x.get("needs_vision") and k not in dis_by and k not in agr_by:
            D.append({"name": x["name"], "score": x["score"], "rub": rubric.get(k), "ev": ev.get(k, "")})
print(f"미검증 {len(D)}건 → Claude 점수 검증")

SYS = """당신은 ESG 평가 감리자입니다. 평가자가 한 지표에 매긴 점수가 제공된 평가지표(채점기준 YN + 메타데이터)와 인용근거에 비춰 올바른지 검증하세요. 평가지표·근거에만 근거하고 Accept/Reject를 엄격히 적용하세요.
verdict: "correct"(점수 맞음)/"wrong"(틀림+올바른점수)/"uncertain"
JSON만: {"results":[{"id":N,"verdict":"correct|wrong|uncertain","correct_score":숫자또는null,"reason":"한 문장"}]}"""


def fmt(i, d):
    r = d["rub"]
    if r:
        m = r.get("meta") or {}
        rub = (f"필수:{r['must']}(미충족 {r['s_mustN']}점)/가산1:{r['add1']}(필수만 {r['s_YN']}점)/"
               f"가산2:{r['add2']}(필수+가산1 {r['s_YYN']}점)/모두충족 {r['s_full']}점\n"
               f"  Intent:{m.get('intent','')}\n  Accept:{m.get('accept','')}\n  Reject:{m.get('reject','')}")
    else:
        rub = "(메타 없음)"
    return f"[id {i}] 지표:{d['name']}\n채점기준:{rub}\n평가자점수:{d['score']}\n인용근거:{d['ev'][:380]}"


def parse(t):
    mt = re.search(r"\{.*\}", t, re.DOTALL)
    s = mt.group(0) if mt else t
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        return json.loads(repair_json(s))


byid = {}
for i in range(0, len(D), 5):
    body = "\n\n".join(fmt(i + j, d) for j, d in enumerate(D[i:i + 5]))
    try:
        msg = client.messages.create(model=MODEL, max_tokens=8000, system=SYS,
                                     messages=[{"role": "user", "content": body}])
        for r in parse(msg.content[0].text).get("results", []):
            byid[r.get("id")] = r
    except Exception as e:
        print(f"  배치 오류 @{i}: {e}")
    time.sleep(1)

out = []
ok = wrong = unc = 0
for i, d in enumerate(D):
    r = byid.get(i, {})
    v = r.get("verdict", "uncertain")
    ok += v == "correct"
    wrong += v == "wrong"
    unc += v == "uncertain"
    out.append({"name": d["name"], "v4_score": d["score"], "verdict": v,
                "claude_score": r.get("correct_score"), "reason": r.get("reason", "")})
print(f"  ✅correct {ok} / ⚠wrong {wrong} / ?uncertain {unc}")
for x in out:
    if x["verdict"] != "correct":
        print(f"   [{x['verdict']}] v4={x['v4_score']} cl={x['claude_score']} | {x['name'][:38]} | {x['reason'][:50]}")
Path(BASE / "output/cross13.json").write_text(json.dumps(out, ensure_ascii=False, indent=2, default=str))
print("저장: cross13.json")

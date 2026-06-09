"""Phase 5: v3·v4가 '동일 점수'로 합의한 지표를 외부 LLM(Claude)로 독립 검증.
합의는 '둘 다 틀림'일 수 있으므로, 평가에 쓴 Gemini가 아닌 Claude에게
평가지표(YN트리+메타데이터)+인용근거만 주고 '합의 점수가 맞는가'를 판정시킨다.
Claude가 wrong이라 보는 건 = v3·v4 공통 오류 후보."""
import sys, re, json, time
from pathlib import Path

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
sys.path.insert(0, str(BASE))
from dotenv import dotenv_values
import openpyxl
import anthropic
from json_repair import repair_json
from core.schemas import load_v4
from verify_results import find_cols

ENV = dotenv_values(str(BASE / ".env"))
client = anthropic.Anthropic(api_key=ENV["ANTHROPIC_API_KEY"])
MODEL = "claude-sonnet-4-20250514"


def norm(s):
    return re.sub(r"[\s\W]", "", str(s or "")).lower()


V3X = next(BASE.glob("output/CJ_v3_v3_*분석결과.xlsx"))
V4X = next(BASE.glob("output/CJ_v4_v4_*분석결과.xlsx"))
V3J = json.loads((BASE / "output" / (V3X.stem + "_verify.json")).read_text())
V4J = json.loads((BASE / "output" / (V4X.stem + "_verify.json")).read_text())
SHEETMAP = [("식품-E", "식품_환경"), ("식품-S", "식품_사회")]

rubric = {}
for sn, inds in load_v4(BASE / "templates/식품_v4.xlsx").items():
    for x in inds:
        rubric[norm(x["name"])] = x


def evidence(path, sheets, keys):
    wb = openpyxl.load_workbook(path, data_only=True)
    m = {}
    for sn in sheets:
        ws = wb[sn]
        name_c, e_c, _ = find_cols(ws)
        for r in range(2, ws.max_row + 1):
            nm = ws.cell(r, name_c).value if name_c else None
            if nm and str(nm).strip() and e_c:
                m[norm(nm)] = str(ws.cell(r, e_c).value or "")
    wb.close()
    return m


ev4 = evidence(V4X, ["식품_환경", "식품_사회"], None)

# 합의: 이름매칭 + 양쪽 정상 + 점수 동일
agree = []
for s3, s4 in SHEETMAP:
    r3 = {norm(x["name"]): x for x in V3J["sheets"].get(s3, {}).get("rows", [])}
    r4 = {norm(x["name"]): x for x in V4J["sheets"].get(s4, {}).get("rows", [])}
    for k in set(r3) & set(r4):
        a, b = r3[k], r4[k]
        if a["cat"] == "정상" and b["cat"] == "정상":
            try:
                sa, sb = float(a["score"]), float(b["score"])
            except (TypeError, ValueError):
                continue
            if sa == sb:
                agree.append({"name": a["name"], "score": sa, "ev": ev4.get(k, ""), "rub": rubric.get(k)})
print(f"합의(동일점수) 지표 {len(agree)}건 → Claude 검증")

SYS = """당신은 ESG 평가 감리자입니다. 한 지표를 두 평가자가 '동일한 점수'로 평가했습니다.
그 합의 점수가 제공된 평가지표(채점기준 YN트리 + 메타데이터: Intent/용어사전/인정형태(Accept)/불인정형태(Reject)/출처요구)와 인용 근거에 비춰 올바른지 독립 검증하세요.
- 오직 제공된 평가지표와 근거에만 근거하세요. 일반지식으로 기준을 뒤집지 마세요. Accept/Reject 패턴은 엄격히 적용하세요.
- verdict: "correct"(합의 점수 맞음) / "wrong"(틀림 — 올바른 점수 제시) / "uncertain"(근거·기준 불충분)
JSON만: {"results":[{"id":N,"verdict":"correct|wrong|uncertain","correct_score":숫자또는null,"reason":"한 문장"}]}"""


def fmt(i, d):
    r = d["rub"]
    if not r:
        rub = "(메타 없음 — 지표 문구로 판단)"
    else:
        m = r.get("meta") or {}
        rub = (f"필수:{r['must']}(미충족 {r['s_mustN']}점)/가산1:{r['add1']}(필수만 {r['s_YN']}점)/"
               f"가산2:{r['add2']}(필수+가산1 {r['s_YYN']}점)/모두충족 {r['s_full']}점(만점)\n"
               f"  Intent:{m.get('intent','')}\n  Accept:{m.get('accept','')}\n  Reject:{m.get('reject','')}")
    return f"[id {i}] 지표: {d['name']}\n채점기준:{rub}\n합의점수: {d['score']}\n인용근거: {d['ev'][:380]}"


def parse(txt):
    mt = re.search(r"\{.*\}", txt, re.DOTALL)
    s = mt.group(0) if mt else txt
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        return json.loads(repair_json(s))


def judge(items, start):
    body = "\n\n".join(fmt(start + i, d) for i, d in enumerate(items))
    msg = client.messages.create(model=MODEL, max_tokens=8000, system=SYS,
                                 messages=[{"role": "user", "content": body}])
    return parse(msg.content[0].text).get("results", [])


byid = {}
for i in range(0, len(agree), 6):
    try:
        for r in judge(agree[i:i + 6], i):
            byid[r.get("id")] = r
    except Exception as e:
        print(f"  배치 오류 @{i}: {e}")
    time.sleep(1)

ok = wrong = unc = unk = 0
rows = []
for i, d in enumerate(agree):
    r = byid.get(i, {})
    v = r.get("verdict", "?")
    ok += v == "correct"
    wrong += v == "wrong"
    unc += v == "uncertain"
    unk += v not in ("correct", "wrong", "uncertain")
    rows.append({"name": d["name"], "consensus": d["score"], "verdict": v,
                 "claude_score": r.get("correct_score"), "reason": r.get("reason", "")})

print(f"\n═══ Claude 외부검증 (합의 {len(agree)}건) ═══")
print(f"  ✅ 합의 확인(correct) {ok} │ ⚠ 합의 반박(wrong) {wrong} │ ? 불확실 {unc} │ 미판정 {unk}")
print(f"\n[⚠ Claude가 v3·v4 합의를 '틀렸다'고 본 건 = 공통오류 후보]")
for x in rows:
    if x["verdict"] == "wrong":
        print(f"  v3=v4={x['consensus']} → Claude={x['claude_score']} | {x['name'][:34]} | {x['reason'][:70]}")
Path(BASE / "output/external_judge_agreements.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=str))
print("\n저장: output/external_judge_agreements.json")

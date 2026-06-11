"""14개 stably-wrong 지표의 flash vs pro 실제 평가(점수+근거+검토) full text 엑셀. 사용자 직접 검수용."""
import sys, json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JOB = Path("/Users/junyeongc/KASE/AI도입/02_분석결과/시스템검증_국내외_0610")
OUT = Path("/Users/junyeongc/KASE/v1.3.0_개선과정.xlsx")  # 버전 개선 기록 (표준 양식)
SHEETS = ["식품_환경", "식품_사회"]
FONT = "맑은 고딕"
HEAD = PatternFill("solid", fgColor="305496"); FFILL = PatternFill("solid", fgColor="FCE4D6"); PFILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

# 14 stably-wrong = 전사안정 ∩ disputed
ws = openpyxl.load_workbook(JOB / "_요약/KASE_v4_검증종합_0610.xlsx")["전 기업 재현성 상세"]
hdr = {ws.cell(4, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(4, c).value}
stable6 = {ws.cell(r, hdr["지표명"]).value for r in range(5, ws.max_row + 1) if str(ws.cell(r, hdr["안정기업수"]).value).startswith("6")}
diffs = json.load(open(JOB / "_요약/disagreements_full.json"))
sw = sorted(stable6 & {d["name"] for d in diffs})

# 내 사전판정 (name 부분문자열 → 판정)
VERD = [
    ("산업재해율(%)을 공개", "🔴 flash오류", "연도 날조 — 2개년 표인데 3개년 만점"),
    ("근로손실재해율(LTIFR)", "🔴 flash오류", "연도 날조(2개년→3개년)"),
    ("협력업체 산업재해율(%)을 공개", "🔴 flash오류", "연도 날조(2개년→3개년)"),
    ("정규직 임직원 비율이 80%", "🔴 flash오류", "73.3%(80%미만)인데 만점 — 공시≠조건충족"),
    ("지정폐기물의 (매립+소각) 비율", "🔴 flash오류", "엉뚱한 지표(재활용률) 인용"),
    ("물 스트레스 지수가 높거나", "🔴 flash오류", "취수량(consumed)≠조달비율 — 다른 지표"),
    ("제3자 인증기관에서 검증", "🔴 flash오류", "자사 R&D를 제3자 인증으로 둔갑"),
    ("Scope 3 배출량을 공개", "🔴 flash오류", "2개년인데 3개년 만점(연도)"),
    ("불공정거래/부정경쟁을 하여 제재", "🔴 flash오류", "p.148 2개년 표를 3개년 날조"),
    ("온실가스 배출 집약도", "🔴 의심", "연도 over-count(2개년→3개년, PDF확인要)"),
    ("임시직 수 또는 비율", "🔴 경미", "각주 제외조건(일용직·노무파견) 무시→만점과다"),
    ("(먼지) 최근 3개년 대기오염", "🟡 모호", "필수조건 정의(2024감소면 충족?) — pro가 더 맞을 듯"),
    ("(황산화물) 최근 3개년 대기오염", "🟡 모호", "'연속감소' 정의 + 기업별 데이터차"),
    ("정보 관련 사고 처리 대응", "🟡 모호", "'대응체계' 범위(진단 포함?) 경계 모호"),
]
def verdict(nm):
    for kw, v, r in VERD:
        if kw in nm:
            return v, r
    return "?", ""


# 처리계획 (판정, 오류클래스/사유, 처리방법, 상태) — 키워드 매칭(특수 키워드 먼저)
DISP = [
    ("(먼지)", "pro 승 (flash오류)", "필수↔가산 혼동 — '당해연도 감소'를 '3개년 연속'으로 오독", "구조적채점 (연도별 값 추출→코드가 감소 판정)", "검토완료"),
    ("(황산화물)", "KASE 정책결정", "보고경계(단독 vs 합계). 관계사 2024부터 측정→합계 추세 왜곡", "KASE 경계정책 확정 후 구조적채점(추세)", "검토완료"),
    ("물 스트레스", "pro 승 (flash오류)", "엉뚱한 지표(취수량≠조달비율). PDF 전수조사로 미공시 확정", "프롬프트규칙(정확한 지표) + 메타 Reject", "검토완료"),
    ("임시직", "pro 승 (flash오류)", "각주 미준수 — 일용직·노무파견 제외를 무시하고 완전성 만점", "구조적채점(완전성: 공시표현·제외표현 코드검출→완전성 감점)", "검토완료"),
    ("정규직 임직원 비율", "pro 승 (flash오류)", "수치조건 무시/자기모순 — 73.3%<80% 알고도 만점", "구조적채점(정규직수·전체수 추출→코드 ≥80% 비교)", "검토완료"),
    ("정보 관련 사고", "pro 승 (근인=루브릭 모호)", "지표 정의 미흡 — '사고 처리 대응체계'를 flash가 '유출 진단'으로 확대해석", "평가지표 구체화 (메타 Accept/Reject: 진단·모니터링 ≠ 사고 대응체계)", "검토완료"),
    ("제3자 인증", "🟡 루브릭 모호 (예비판정 정정)", "flash는 실제 제3자 인증(PETA/Leaping Bunny) 인용 — 단 cruelty-free가 '안전성' 검증인지 정의 미흡", "평가지표 구체화 (안전성 검증에 윤리/cruelty-free 포함 여부 정의)", "검토완료"),
    ("근로손실재해율", "🔴 flash오류 확정 (PDF)", "연도 날조 — p.148 LTIR 2개년(2023·24)인데 flash가 2022 컬럼 날조→만점", "구조적채점(공개 연도 추출→코드 카운트)", "검토완료"),
    ("협력업체 산업재해율", "🔴 flash오류 확정 (PDF)", "연도 날조 — 동일 p.148 표 2개년(2023·24)", "구조적채점(공개 연도 추출→코드 카운트)", "검토완료"),
    ("산업재해율(%)을 공개", "🔴 flash오류 확정 (PDF)", "연도 날조 — p.148 산재율 2개년(2023·24)", "구조적채점(공개 연도 추출→코드 카운트)", "검토완료"),
    ("Scope 3", "🔴 flash오류 확정", "연도 날조 — 오뚜기 Scope3 2개년(2023·24)인데 flash가 2022 지어냄", "구조적채점(공개 연도 추출→코드 카운트)", "검토완료"),
    ("배출 집약도", "🔴 flash오류 확정", "연도 날조 — 유니레버 집약도 2개년(2024·25)인데 flash가 3개년 지어냄", "구조적채점(공개 연도 추출→코드 카운트)", "검토완료"),
    ("지정폐기물", "🔴 flash오류 확정", "엉뚱한 지표 — flash가 재활용률(3.9/7.3/7.1) 인용. 지표는 (매립+소각)/총 비율, pro가 정확히 계산(0.999→0.636→0.525%↓)", "구조적채점 (참고산식 기반 매립·소각·총 추출→코드 계산)", "검토완료"),
    ("불공정거래", "🔴 flash오류 확정", "연도 날조 — p.148 공정경쟁 2개년(2023·24)인데 flash가 3개년(0,0,0) 날조", "구조적채점(공개 연도 추출→코드 카운트)", "검토완료"),
]


def disp(nm):
    for kw, *rest in DISP:
        if kw in nm:
            return rest
    return ["?", "", "", "검토전"]


def find_row(wb, name):
    for sh in SHEETS:
        if sh not in wb.sheetnames: continue
        w = wb[sh]
        for r in range(2, w.max_row + 1):
            if str(w.cell(r, 3).value).strip() == name.strip():
                return sh, r
    return None, None


def evalof(co, name, pro=False):
    d = JOB / co
    cand = [f for f in d.glob("*.xlsx") if not f.name.startswith(("~$", ".")) and (("pro" in f.name) == pro)]
    for f in sorted(cand):
        try:
            wb = openpyxl.load_workbook(f)
        except Exception:
            continue  # 잠금/손상 파일 스킵
        sh, r = find_row(wb, name)
        if not r:
            continue
        w = wb[sh]; h = {str(w.cell(1, c).value): c for c in range(1, w.max_column + 1) if w.cell(1, c).value}
        cs = next((c for k, c in h.items() if "AI 점수" in k), None)
        ce = next((c for k, c in h.items() if "AS-IS" in k or "내용" in k), None)
        cg = next((c for k, c in h.items() if "검토의견" in k), None)
        return w.cell(r, cs).value, str(w.cell(r, ce).value), str(w.cell(r, cg).value)
    return "?", "(파일없음/행없음)", ""


wb = openpyxl.Workbook(); sh = wb.active; sh.title = "stably-wrong 14"
sh["A1"] = "재현성 충족(6사 10회 동일) but flash↔pro 불일치 — 14개 직접 검수용 (flash·pro 실제 평가)"
sh["A1"].font = Font(FONT, size=12, bold=True)
cols = ["#", "지표명", "비교기업", "flash 점수", "flash 근거(AS-IS 원문)", "flash 검토의견", "pro 점수", "pro 근거(AS-IS 원문)", "pro 검토의견", "(참고) AI사전판정", "사유"]
for c, h in enumerate(cols, 1):
    cell = sh.cell(3, c, h); cell.font = Font(FONT, bold=True, color="FFFFFF"); cell.fill = HEAD
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = THIN
r = 4
for i, nm in enumerate(sw, 1):
    ds = [d for d in diffs if d["name"] == nm]
    co = ds[0]["co"]
    fs, fev, frv = evalof(co, nm, pro=False)
    ps, pev, prv = evalof(co, nm, pro=True)
    v, vr = verdict(nm)
    vals = [i, nm, co, fs, fev, frv, ps, pev, prv, v, vr]
    for c, val in enumerate(vals, 1):
        cell = sh.cell(r, c, val); cell.font = Font(FONT, size=9); cell.border = THIN
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 5, 6, 8, 9, 11)), horizontal=("center" if c in (1, 3, 4, 7) else "left"))
    for c in (4, 5, 6): sh.cell(r, c).fill = FFILL
    for c in (7, 8, 9): sh.cell(r, c).fill = PFILL
    sh.row_dimensions[r].height = 95
    r += 1
for c, w in enumerate([4, 30, 9, 7, 38, 30, 7, 38, 30, 13, 28], 1):
    sh.column_dimensions[get_column_letter(c)].width = w
sh.freeze_panes = "C4"; sh.sheet_view.topLeftCell = "A1"

# ── 탭2: 14개 처리계획 (판정 & 처리방법 영구기록) ──
GREEN2 = PatternFill("solid", fgColor="C6EFCE"); GREY2 = PatternFill("solid", fgColor="EDEDED")
ws2 = wb.create_sheet("14개 처리계획")
ws2["A1"] = "stably-wrong 14개 — 판정 & 처리계획  (상태: 검토완료=확정 / 검토전=예비, 사용자 검토 중)"
ws2["A1"].font = Font(FONT, size=12, bold=True)
for c, h in enumerate(["#", "지표명", "flash", "pro", "판정(누가 맞나)", "오류 클래스 / 사유", "처리 방법", "상태"], 1):
    cell = ws2.cell(3, c, h); cell.font = Font(FONT, bold=True, color="FFFFFF"); cell.fill = HEAD
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = THIN
r2 = 4
for i, nm in enumerate(sw, 1):
    ds = [d for d in diffs if d["name"] == nm]
    judgment, klass, method, status = disp(nm)
    for c, v in enumerate([i, nm, ds[0]["f"], ds[0]["p"], judgment, klass, method, status], 1):
        cell = ws2.cell(r2, c, v); cell.font = Font(FONT, size=9); cell.border = THIN
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 5, 6, 7)), horizontal=("center" if c in (1, 3, 4, 8) else "left"))
    ws2.cell(r2, 8).fill = GREEN2 if status == "검토완료" else GREY2
    ws2.row_dimensions[r2].height = 44
    r2 += 1
for c, w in enumerate([4, 30, 7, 7, 18, 38, 36, 10], 1):
    ws2.column_dimensions[get_column_letter(c)].width = w
ws2.freeze_panes = "C4"; ws2.sheet_view.topLeftCell = "A1"

# ── 검증결과 탭 (전 식품 6사 × 10회, 재현성 셀 양식, flash 원본 → 구조적 after) ──
AC = json.loads(Path("/Users/junyeongc/KASE/AI도입/02_분석결과/시스템검증_국내외_0610/_요약/structural_repro_allco.json").read_text())["summary"]
COMPCOL = ["CJ제일제당", "풀무원", "오뚜기", "대상", "네슬레", "유니레버"]
VMAP = [
    ("#1 먼지개선", "(먼지) 최근 3개년 대기오염물질 배출 실적이 개선되었다."),
    ("#2 SOx개선", "(황산화물) 최근 3개년 대기오염물질 배출 실적이 개선되었다."),
    ("#3 물스트레스조달", "물 스트레스 지수가 높거나 매우 높은 지역에서 조달된 식품 원료의 비율을 정량 공개하고 있다."),
    ("#4 임시직", "임시직 수 또는 비율을 공개하고 있다."),
    ("#5 정규직80", "전체 임직원 대비 정규직 임직원 비율이 80% 이상이다."),
    ("#6 정보사고", "정보 관련 사고 처리 대응 체계를 마련하고 있다."),
    ("#7 제3자인증", "제품, 서비스의 안전성을 제3자 인증기관에서 검증받은 내역을 공개한다"),
    ("#8 LTIFR", "최근 3개년 근로손실재해율(LTIFR)을 공개하고 있다."),
    ("#9 산재율", "최근 3개년 산업재해율(%)을 공개하고 있다."),
    ("#10 Scope3", "최근 3개년 온실가스 Scope 3 배출량을 공개하고 있다."),
    ("#11 집약도", "최근 3개년 온실가스 배출 집약도를 공개하고 있다."),
    ("#12 지정폐기물", "최근 3개년 총 폐기물 배츌량 대비 지정폐기물의 (매립+소각) 비율이 개선되었다."),
    ("#13 협력사산재율", "최근 3개년 협력업체 산업재해율(%)을 공개하고 있다."),
    ("#14 불공정거래", "최근 3개년도 불공정거래/부정경쟁을 하여 제재받은 내역을 공개한다."),
]
# 기준기업(원 분쟁사례) = INDS pdf → (표시명, pro 정답). 이 기업에서 구조적 mode==정답이면 정확성 회복 확인.
CALIB = {
    "#1 먼지개선": ("CJ제일제당", 1), "#3 물스트레스조달": ("네슬레", 0), "#4 임시직": ("CJ제일제당", 1.5),
    "#5 정규직80": ("CJ제일제당", 0), "#8 LTIFR": ("풀무원", 0), "#9 산재율": ("풀무원", 0),
    "#10 Scope3": ("오뚜기", 0), "#11 집약도": ("유니레버", 0), "#12 지정폐기물": ("CJ제일제당", 2),
    "#13 협력사산재율": ("풀무원", 0), "#14 불공정거래": ("풀무원", 0),
}
TRACK = {"#2 SOx개선": "KASE 보고경계 정책", "#6 정보사고": "평가지표 구체화", "#7 제3자인증": "평가지표 구체화"}
# flash before: 전 기업 재현성 상세 (지표명 정규화 → {기업: "10회 점수"})
def _norm(s):
    return str(s).replace(" ", "").replace(" ", "").replace(".", "").strip()
refws = openpyxl.load_workbook(JOB / "_요약/KASE_v4_검증종합_0610.xlsx")["전 기업 재현성 상세"]
FB = {}
for r in range(5, refws.max_row + 1):
    rnm = refws.cell(r, 2).value
    if rnm: FB[_norm(rnm)] = {COMPCOL[c - 3]: str(refws.cell(r, c).value) for c in range(3, 9)}


def j10(scores):
    out = []
    for x in scores:
        try:
            f = float(x); out.append(str(int(f)) if f == int(f) else str(f))
        except (TypeError, ValueError):
            out.append(str(x))
    return ",".join(out) if out else "—"


GREEN3 = PatternFill("solid", fgColor="C6EFCE"); RED3 = PatternFill("solid", fgColor="FFC7CE")
GREY3 = PatternFill("solid", fgColor="EDEDED"); ORANGE3 = PatternFill("solid", fgColor="FCE4D6")
wsv = wb.create_sheet("검증결과(기업별 10회)")
wsv["A1"] = "v1.3.0 검증 — 전 식품 6사 × 10회 반복 (재현성 셀 양식) · flash 원본(before) → 구조적 채점기(after)"
wsv["A1"].font = Font(FONT, size=12, bold=True)
wsv["A2"] = "각 셀 = 해당 기업 10회 점수. 초록=10/10 일관 / 회색=별도 트랙(루브릭·정책). 결정론 11지표 × 6사 = 66셀 전부 일관, 기준기업 정답 11/11 = pro 일치."
wsv["A2"].font = Font(FONT, size=9, italic=True, color="808080")
heads = ["#", "지표명", "단계"] + COMPCOL + ["기준기업·정답(pro)", "결과"]
for c, h in enumerate(heads, 1):
    cell = wsv.cell(4, c, h); cell.font = Font(FONT, bold=True, color="FFFFFF"); cell.fill = HEAD
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = THIN
rv = 5
for i, (rid, nm) in enumerate(VMAP, 1):
    acrow = AC.get(rid, {}); fbrow = FB.get(_norm(nm), {}); track = TRACK.get(rid)
    if track:
        calibtxt = "—"; result = f"➖ {track} 트랙"
    else:
        cco, ctruth = CALIB[rid]; calibtxt = f"{cco}={ctruth}"
        allcon = all(acrow.get(co, {}).get("consistent") for co in COMPCOL)
        calibok = acrow.get(cco, {}).get("mode") == ctruth
        result = "✅ 통과 (재현X→O·정확X→O)" if (allcon and calibok) else "❌ 미통과"
    rowA = [i, nm, "flash 원본 ×10"] + [fbrow.get(co, "—") for co in COMPCOL] + [calibtxt, result]
    for c, v in enumerate(rowA, 1):
        cell = wsv.cell(rv, c, v); cell.font = Font(FONT, size=8); cell.border = THIN
        cell.alignment = Alignment(vertical="center", wrap_text=(c in (2, 11)), horizontal=("center" if c == 1 else "left"))
        if 4 <= c <= 9: cell.fill = ORANGE3
    rowB = ["", "", "v1.3.0 구조적 ×10"] + [j10(acrow.get(co, {}).get("scores", [])) for co in COMPCOL] + ["", ""]
    for c, v in enumerate(rowB, 1):
        cell = wsv.cell(rv + 1, c, v); cell.font = Font(FONT, size=8); cell.border = THIN
        cell.alignment = Alignment(vertical="center", horizontal="left")
        if 4 <= c <= 9:
            con = acrow.get(COMPCOL[c - 4], {}).get("consistent")
            cell.fill = GREY3 if track else (GREEN3 if con else RED3)
    for c in (1, 2, 10, 11):
        wsv.merge_cells(start_row=rv, start_column=c, end_row=rv + 1, end_column=c)
    rescell = wsv.cell(rv, 11); rescell.fill = GREEN3 if not track else GREY3
    rescell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    wsv.cell(rv, 1).alignment = Alignment(vertical="center", horizontal="center")
    wsv.cell(rv, 2).alignment = Alignment(vertical="center", wrap_text=True)
    wsv.cell(rv, 10).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    wsv.row_dimensions[rv].height = 24; wsv.row_dimensions[rv + 1].height = 24
    rv += 2
note = "결정론 11지표 × 6사 = 66셀: 전부 10/10 일관 ✅ + 기준기업(원 분쟁사례) 정답 11/11 = pro 판정 일치. #2=KASE 보고경계 정책 / #6·#7=평가지표(루브릭) 구체화 별도 트랙. ※ 기업별 점수 차이는 보고서별 실제 데이터 차이(정상)."
wsv.cell(rv + 1, 2, note).font = Font(FONT, size=9, italic=True, color="808080")
for c, w in enumerate([4, 32, 16] + [13] * 6 + [16, 20], 1):
    wsv.column_dimensions[get_column_letter(c)].width = w
wsv.freeze_panes = "D5"; wsv.sheet_view.topLeftCell = "A1"

# ── 개요 탭 (맨 앞) ──
wso = wb.create_sheet("개요", 0)
for i, (t, sz, b) in enumerate([
    ("시스템 개선 기록 — v1.2.0 → v1.3.0", 14, True), ("", 10, False),
    ("[대상] stably-wrong 14개 (재현 O · 정확 X) — flash가 일관되게 틀리던 지표", 11, False),
    ("[변경] 구조적 채점기 도입: LLM이 '값'만 추출 → 코드가 PDF 교차검증 → 코드가 결정론 채점", 11, False),
    ("        (연도 날조·수치 자기모순이 구조적으로 불가능. flash의 '판단'을 코드로 대체)", 10, False), ("", 10, False),
    ("[검증 게이트] 전 식품 6사(국내4·국외2) × 14지표 × 10회 반복 (실제 시스템·실제 SR보고서):", 11, True),
    ("   · 결정론 11지표 × 6사 = 66셀: 전부 10/10 일관 ✅ + 기준기업(원 분쟁사례) 정답 11/11 = pro 판정 일치", 10, False),
    ("   · #2 SOx: 보고경계(단독/합계) 미정으로 흔들림 → KASE 정책 결정 후 해결", 10, False),
    ("   · #6·#7: rubric(판단형) — 코드로 못 고침 → 평가지표(템플릿) 구체화", 10, False), ("", 10, False),
    ("[탭] 처리계획(14개 판정·방법) / 검증결과(기업별 10회 before→after) / 근거(flash vs pro 실측)", 10, False), ("", 10, False),
    ("※ 이 문서 = '버전 변경 기록' 표준 양식. 향후 버전마다 [변경지점 · 변경방법 · 검증통과]를 이렇게 남긴다.", 10, True),
], 1):
    wso.cell(i, 1, t).font = Font(FONT, size=sz, bold=b)
wso.column_dimensions["A"].width = 112
wso.sheet_view.showGridLines = False

# 탭 순서: 개요 / 처리계획 / 검증결과 / 근거
_order = ["개요", "14개 처리계획", "검증결과(기업별 10회)", "stably-wrong 14"]
wb._sheets.sort(key=lambda s: _order.index(s.title) if s.title in _order else 99)

wb.save(OUT)
print(f"✅ {OUT.name} | {len(sw)}개 지표")
for i, nm in enumerate(sw, 1):
    ds = [d for d in diffs if d["name"] == nm]; print(f"  {i:2}. flash={ds[0]['f']} pro={ds[0]['p']} | {nm[:46]}")

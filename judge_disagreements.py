"""Phase 2-② / Phase 4: v3·v4 점수 불일치 지표를 LLM judge로 독립 판정.
이름매칭된 불일치(양쪽 정상·점수 다름)만 대상. 각 지표를 A(v3)/B(v4) 익명화하여
채점기준(YN트리)+각자 인용근거에 비춰 올바른 점수와 누가 맞는지 판정 → v4 정확도 집계.
(근거 기반 판정 — 전체 SR 재독은 아님. 출처 grounding은 이미 ~100% 확인됨)"""
import sys, re, json, time
from pathlib import Path

BASE = Path("/Users/junyeongc/KASE/AI도입/KASE_ESG_Analyzer")
sys.path.insert(0, str(BASE))
from dotenv import dotenv_values
import openpyxl
from google import genai
from google.genai import types
from json_repair import repair_json
from core.schemas import load_v4
from verify_results import find_cols


def norm(s):
    return re.sub(r"[\s\W]", "", str(s or "")).lower()


ENV = dotenv_values(str(BASE / ".env"))
client = genai.Client(api_key=ENV["GOOGLE_API_KEY"])

V3X = next(BASE.glob("output/CJ_v3_v3_*분석결과.xlsx"))
V4X = next(BASE.glob("output/CJ_v4_v4_*분석결과.xlsx"))
V3J = json.loads((BASE / "output" / (V3X.stem + "_verify.json")).read_text())
V4J = json.loads((BASE / "output" / (V4X.stem + "_verify.json")).read_text())
SHEETMAP = [("식품-E", "식품_환경"), ("식품-S", "식품_사회")]

rubric = {}
for sn, inds in load_v4(BASE / "templates/식품_v4.xlsx").items():
    for x in inds:
        rubric[norm(x["name"])] = x


def econtent_map(xlsx):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    m = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        name_c, e_c, _ = find_cols(ws)
        if not (name_c and e_c):
            continue
        for r in range(2, ws.max_row + 1):
            nm = ws.cell(r, name_c).value
            if nm and str(nm).strip():
                m[norm(nm)] = str(ws.cell(r, e_c).value or "")
    wb.close()
    return m


e3, e4 = econtent_map(V3X), econtent_map(V4X)

disag = []
for s3, s4 in SHEETMAP:
    r3 = {norm(x["name"]): x for x in V3J["sheets"].get(s3, {}).get("rows", [])}
    r4 = {norm(x["name"]): x for x in V4J["sheets"].get(s4, {}).get("rows", [])}
    for k in set(r3) & set(r4):
        a, b = r3[k], r4[k]
        if a["cat"] == "정상" and b["cat"] == "정상":
            try:
                sa, sb = float(a["score"]), float(b["score"])
            except (TypeError, ValueError):
                continue
            if sa != sb:
                disag.append({"name": a["name"], "v3": sa, "v4": sb,
                              "e3": e3.get(k, ""), "e4": e4.get(k, ""), "rub": rubric.get(k)})
print(f"이름매칭 불일치 {len(disag)}건 (양쪽 정상·점수 다름)")

SYS = """당신은 ESG 평가 감리자입니다. 한 지표를 두 평가자 A·B가 서로 다른 점수로 평가했습니다.

★ 판정 원칙: **오직 아래 제공된 '평가지표'(채점기준 YN트리 + 메타데이터: Intent/용어사전/인정형태(Accept)/불인정형태(Reject)/출처요구)에만 근거해 판정하라.**
- 당신의 일반지식으로 기준을 보완하거나 뒤집지 마라.
- 특히 Accept/Reject 패턴을 엄격히 적용하라: Reject에 명시된 형태의 근거는 (일반적으로는 타당해 보여도) **인정 불가**다.
- 인용 근거가 메타데이터 기준을 정확히 충족하는지만 따져라.

verdict: "A"(A 점수가 평가지표상 맞음) / "B"(B가 맞음) / "both_ok" / "neither"

★ 추가: 만약 *이 지표의 평가기준(특히 Accept/Reject) 자체가 일반적 ESG 도메인 상식과 충돌*한다고 보이면(예: 도메인에선 통용되는 형태를 평가지표가 Reject로 막는 경우) conflict=true와 이유를 적어라. 단 verdict/점수는 어디까지나 평가지표 기준으로 판정하라.

JSON만: {"results":[{"id":N,"correct_score":숫자,"verdict":"A|B|both_ok|neither","reason":"한 문장","conflict":true/false,"conflict_note":"충돌 시 설명(평가지표 vs 도메인상식), 없으면 빈문자열"}]}"""


def fmt(i, d):
    r = d["rub"]
    if not r:
        return (f"[id {i}] 지표: {d['name']}\n(평가지표 메타 없음 — 지표 문구·근거로만 판정)\n"
                f"A 근거: {d['e3'][:350]} → A점수={d['v3']}\nB 근거: {d['e4'][:350]} → B점수={d['v4']}")
    m = r.get("meta") or {}
    yn = (f"필수:{r['must']}(미충족 {r['s_mustN']}점) / 가산1:{r['add1']}(필수만 {r['s_YN']}점) / "
          f"가산2:{r['add2']}(필수+가산1 {r['s_YYN']}점) / 모두충족 {r['s_full']}점(만점)")
    meta = (f"  - 지표의도(Intent): {m.get('intent','')}\n"
            f"  - 용어사전(Term): {m.get('terms','')}\n"
            f"  - 인정형태(Accept): {m.get('accept','')}\n"
            f"  - 불인정형태(Reject): {m.get('reject','')}\n"
            f"  - 출처요구(Citation): {m.get('citation','')}")
    return (f"[id {i}] 지표: {d['name']}\n[채점기준 YN] {yn}\n[메타데이터]\n{meta}\n"
            f"A 근거: {d['e3'][:350]} → A점수={d['v3']}\n"
            f"B 근거: {d['e4'][:350]} → B점수={d['v4']}")


def judge_batch(items, start):
    body = "\n\n".join(fmt(start + i, d) for i, d in enumerate(items))
    resp = client.models.generate_content(
        model="gemini-2.5-flash",
        contents=[types.Content(parts=[types.Part.from_text(text=SYS + "\n\n" + body)])],
        config=types.GenerateContentConfig(max_output_tokens=16000, temperature=0.1, response_mime_type="application/json"))
    txt = resp.text
    try:
        data = json.loads(txt)
    except json.JSONDecodeError:
        data = json.loads(repair_json(txt))
    return data.get("results", [])


byid = {}
for i in range(0, len(disag), 6):
    chunk = disag[i:i + 6]
    try:
        for r in judge_batch(chunk, i):
            byid[r.get("id")] = r
    except Exception as e:
        print(f"  judge 배치 오류 @{i}: {e}")
    time.sleep(1)

v4r = v3r = both = neither = unk = 0
rows = []
for i, d in enumerate(disag):
    r = byid.get(i, {})
    v = r.get("verdict", "?")
    if v == "B":
        v4r += 1
    elif v == "A":
        v3r += 1
    elif v == "both_ok":
        both += 1
    elif v == "neither":
        neither += 1
    else:
        unk += 1
    has_meta = bool(d["rub"] and any((d["rub"].get("meta") or {}).values()))
    rows.append({"name": d["name"], "v3": d["v3"], "v4": d["v4"], "has_meta": has_meta,
                 "verdict": v, "correct": r.get("correct_score"), "reason": r.get("reason", ""),
                 "conflict": bool(r.get("conflict")), "conflict_note": r.get("conflict_note", "")})

print(f"\n═══ 판정 결과 (불일치 {len(disag)}건, 평가지표 엄격 적용) ═══")
print(f"  [전체]        v4 맞음 {v4r} │ v3 맞음 {v3r} │ 둘 다 타당 {both} │ 둘 다 틀림 {neither} │ 미판정 {unk}")
mrows = [x for x in rows if x["has_meta"]]
mv4 = sum(1 for x in mrows if x["verdict"] == "B")
mv3 = sum(1 for x in mrows if x["verdict"] == "A")
print(f"  [메타보유 {len(mrows)}건] v4 맞음 {mv4} │ v3 맞음 {mv3} │ 나머지 {len(mrows) - mv4 - mv3}")
print(f"\n[상세] (📋=메타보유)")
for x in rows:
    tag = {"B": "✅v4", "A": "❌v3", "both_ok": "～둘다", "neither": "⚠둘다틀림"}.get(x["verdict"], "?")
    m = "📋" if x["has_meta"] else "  "
    print(f"  {m}{tag} v3={x['v3']} v4={x['v4']} 정답={x.get('correct')} | {x['name'][:30]} | {x['reason'][:58]}")

conf = [x for x in rows if x["conflict"]]
print(f"\n═══ ⚠ 평가지표 ↔ 일반지식 '충돌' {len(conf)}건 ═══")
for x in conf:
    print(f"  • {x['name'][:30]} | {x['conflict_note'][:100]}")

Path(BASE / "output/judge_result.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2, default=str))
print("\n저장: output/judge_result.json")
